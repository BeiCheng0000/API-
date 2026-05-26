
"""
API自动化测试平台 - Web应用

提供可视浏览器操作功能，包括：
- 测试用例管理（CRUD）
- 接口调试（支持断言）
- 定时任务调度（APScheduler + Cron，支持持久化）
- 测试执行与结果查看
- 项目管理
- 统计面板

启动方式：
    python run_web_app.py          # 生产模式（Waitress）
    python run_web_app.py --dev    # 开发模式（Flask）
    python run_web_app.py --port 8080  # 自定义端口

注意：此文件不应通过 pytest 运行，已设置 __test__ = False

文件结构导航：
    ├─ 全局初始化（1-108行）
    │   ├─ Flask 应用创建与配置
    │   ├─ 数据库连接与自动迁移
    │   └─ 全局 HTTP Session（复用TCP连接）
    │
    ├─ 定时任务管理（109-300行）
    │   ├─ _save_scheduler_jobs()  - 保存定时任务到数据库
    │   ├─ _extract_cron_expression() - 从CronTrigger提取cron表达式
    │   ├─ _load_scheduler_jobs() - 从数据库恢复定时任务
    │   └─ APScheduler 初始化与启动
    │
    ├─ 辅助函数（300-1150行）
    │   ├─ 占位符替换（${timestamp} 等）
    │   ├─ 断言处理与比较
    │   ├─ 统计数据记录
    │   └─ 数据文件读写
    │
    ├─ 页面路由（1154-1200行）
    │   ├─ GET /            - 主页面
    │   └─ GET /api/top_stats - 顶部统计
    │
    ├─ 项目管理 API（1200-1850行）
    │   ├─ POST /projects/add          - 添加项目
    │   ├─ POST /projects/delete/<name> - 删除项目
    │   ├─ POST /projects/update       - 更新项目
    │   ├─ GET  /projects/list         - 项目列表
    │   ├─ 模块管理（增删改查）
    │   └─ API管理（增删改查）
    │
    ├─ 测试数据 API（1850-1990行）
    │   ├─ POST /api/add              - 添加测试用例
    │   ├─ GET  /api/delete/<name>/<idx> - 删除用例
    │   └─ POST /api/edit/<name>/<idx>   - 编辑用例
    │
    ├─ 接口调试 API（1989-2000行）
    │   └─ POST /api/debug            - 调试接口
    │
    ├─ 定时任务 API（1999-2160行）
    │   ├─ GET  /scheduler/list       - 任务列表
    │   ├─ POST /scheduler/add        - 添加任务
    │   ├─ POST /scheduler/update     - 更新任务
    │   └─ POST /scheduler/delete/<id> - 删除任务
    │
    ├─ 统计 API（2158-2500行）
    │   ├─ GET  /statistics/projects  - 项目统计
    │   ├─ GET  /statistics/modules   - 模块统计
    │   ├─ GET  /statistics/list      - 统计列表
    │   ├─ GET  /statistics/export    - 导出统计
    │   └─ GET  /statistics/detail/<id> - 统计详情
    │
    ├─ 环境与变量管理 API（2500-2730行）
    │   ├─ 环境切换与保存
    │   └─ 变量增删改查
    │
    ├─ 测试执行 API（2726-2890行）
    │   ├─ POST /test/execute_module/<proj>/<mod> - 执行模块
    │   └─ POST /test/execute/<proj>/<mod>/<id>   - 执行单个
    │
    └─ 数据备份与恢复 API（2885-3004行）
        ├─ POST /statistics/restore     - 恢复统计
        ├─ GET  /statistics/backup/list - 备份列表
        └─ GET  /statistics/download    - 下载备份
"""

# 告诉Pytest忽略这个文件
__test__ = False

import os
import json
import time
import hashlib
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional

# 在所有业务模块导入之前加载 .env 环境变量
# 这样数据库密码等敏感信息可以通过 .env 文件配置
from dotenv import load_dotenv
load_dotenv()

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from flask_compress import Compress

from api.base_api import BaseAPI
from common.yaml_handler import YamlHandler
from common.logger_handler import logger
from common.config_handler import env_config
from common.db_handler import MySQLHandler
from common.email_handler import email_handler
from utils.cache_bypass_new import _get_projects_directly
from utils.session_manager import session_manager
from utils.get_projects_fixed_new import get_projects as get_projects_fixed

# 缓存相关
from functools import wraps
import threading

# 简单的内存缓存实现
class SimpleCache:
    def __init__(self, default_timeout=300):  # 默认缓存5分钟
        self.cache = {}
        self.default_timeout = default_timeout
        self.lock = threading.Lock()

    def get(self, key):
        with self.lock:
            item = self.cache.get(key)
            if item is None:
                return None
            if time.time() > item['expires']:
                del self.cache[key]
                return None
            return item['value']

    def set(self, key, value, timeout=None):
        if timeout is None:
            timeout = self.default_timeout
        with self.lock:
            self.cache[key] = {
                'value': value,
                'expires': time.time() + timeout
            }

    def clear(self):
        with self.lock:
            self.cache.clear()

# 创建缓存实例
projects_cache = SimpleCache(default_timeout=60)  # 项目数据缓存1分钟
stats_cache = SimpleCache(default_timeout=30)  # 统计数据缓存30秒
scheduler_cache = SimpleCache(default_timeout=15)  # 定时任务缓存15秒

# 清除缓存函数
def clear_project_cache():
    """清除项目相关的缓存"""
    logger.info("清除项目缓存")
    projects_cache.clear()
    stats_cache.clear()
    scheduler_cache.clear()

# 缓存装饰器
def cache_result(cache_key, cache_instance, timeout=None):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # 生成缓存键
            key_parts = [cache_key]
            if args:
                key_parts.extend(str(arg) for arg in args)
            if kwargs:
                key_parts.extend(f"{k}={v}" for k, v in sorted(kwargs.items()))
            cache_key_full = hashlib.md5("|".join(key_parts).encode()).hexdigest()

            # 尝试从缓存获取
            result = cache_instance.get(cache_key_full)
            if result is not None:
                logger.debug(f"缓存命中: {cache_key_full}")
                return result

            # 缓存未命中，执行函数
            logger.debug(f"缓存未命中: {cache_key_full}")
            result = func(*args, **kwargs)

            # 存入缓存
            cache_instance.set(cache_key_full, result, timeout)
            return result
        return wrapper
    return decorator

# 创建Flask应用
app = Flask(__name__)
app.secret_key = 'api_automation_platform_secret_key_2024'  # ⚠️ 生产环境请修改为随机字符串，建议通过环境变量 FLASK_SECRET_KEY 设置

# 启用GZIP压缩
Compress(app)

# 全局HTTP Session，复用TCP连接，避免每次请求都重新建立连接（DNS解析+TCP握手+TLS握手约300ms）
import requests as _requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# 配置重试策略
retry_strategy = Retry(
    total=2,
    backoff_factor=0.5,
    status_forcelist=[429, 500, 502, 503, 504]
)

# 配置连接池
adapter = HTTPAdapter(
    max_retries=retry_strategy,
    pool_connections=20,
    pool_maxsize=100,
    pool_block=False
)

# 创建HTTP会话并配置连接池
_http_session = _requests.Session()
_http_session.mount("http://", adapter)
_http_session.mount("https://", adapter)

# 数据库处理器（用于统计数据持久化）
db_handler = None
try:
    db_handler = MySQLHandler()
    db_handler.connect()
    logger.info("数据库连接成功，统计数据将保存到数据库")

    # 自动迁移：为已有数据库添加 timestamp 和 error 字段
    try:
        columns = db_handler.query(
            "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'test_statistics'"
        )
        column_names = [col['COLUMN_NAME'] for col in columns] if columns else []

        if 'timestamp' not in column_names:
            db_handler.execute(
                "ALTER TABLE test_statistics ADD COLUMN `timestamp` VARCHAR(30) COMMENT '记录时间戳(前端展示用)' AFTER `response_body`"
            )
            db_handler.execute(
                "UPDATE test_statistics SET timestamp = DATE_FORMAT(created_at, '%%Y-%%m-%%d %%H:%%i:%%s') WHERE timestamp IS NULL OR timestamp = ''"
            )
            logger.info("自动迁移：已添加 timestamp 字段并同步历史数据")

        if 'error' not in column_names:
            db_handler.execute(
                "ALTER TABLE test_statistics ADD COLUMN `error` TEXT COMMENT '错误信息' AFTER `timestamp`"
            )
            logger.info("自动迁移：已添加 error 字段")
    except Exception as migrate_err:
        logger.warning(f"自动迁移检查失败（新数据库无需迁移）: {migrate_err}")

    # 自动迁移：为projects表添加current_env字段
    try:
        columns = db_handler.query(
            "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'projects'"
        )
        column_names = [col['COLUMN_NAME'] for col in columns] if columns else []

        if 'current_env' not in column_names:
            db_handler.execute(
                "ALTER TABLE projects ADD COLUMN `current_env` VARCHAR(255) DEFAULT '' COMMENT '当前激活的环境名称'"
            )
            logger.info("自动迁移：已为projects表添加current_env字段")
    except Exception as migrate_err:
        logger.warning(f"自动迁移projects表失败（新数据库无需迁移）: {migrate_err}")

    # 启动时预加载统计数据
    try:
        # logger.info("=" * 50)
        # logger.info("启动预加载：开始从数据库获取统计数据...")
        db_handler._ensure_connection()
        preload_count = db_handler.query("SELECT COUNT(*) as cnt FROM test_statistics")
        total_count = preload_count[0]['cnt'] if preload_count else 0
        # logger.info(f"启动预加载：数据库中共有 {total_count} 条统计数据")
        if total_count > 0:
            sample_data = db_handler.query("SELECT id, project, module, case_name, method, status_code FROM test_statistics ORDER BY id DESC LIMIT 5")
            # logger.info(f"启动预加载：最近5条数据预览:")
            # for row in sample_data:
            #     logger.info(f"  - ID:{row.get('id')} | 项目:{row.get('project','')} | 模块:{row.get('module','')} | 用例:{row.get('case_name','')} | 方法:{row.get('method','')} | 状态码:{row.get('status_code','')}")
        # logger.info("启动预加载：统计数据加载完成")
        # logger.info("=" * 50)
    except Exception as preload_err:
        logger.warning(f"启动预加载统计数据失败: {preload_err}")

except Exception as e:
    logger.warning(f"数据库连接失败: {e}，统计数据将保存到文件")

# 项目根目录
BASE_DIR = Path(__file__).parent
DATA_DIR = os.path.join(BASE_DIR, "data")

# 统计数据文件路径（仅用于备份/恢复功能的路径校验）
STATISTICS_FILE = os.path.join(DATA_DIR, "statistics.json")


def _save_scheduler_jobs():
    """将当前定时任务保存到数据库"""
    try:
        if not db_handler:
            logger.error("db_handler 为 None，无法保存定时任务")
            return

        db_handler._ensure_connection()

        # 先读取当前数据库中的排序信息，避免保存时丢失
        existing_orders = {}
        try:
            existing_rows = db_handler.query("SELECT id, sort_order FROM scheduler_jobs")
            for row in existing_rows:
                existing_orders[row['id']] = row.get('sort_order', 0) or 0
        except Exception:
            pass

        # 清空旧数据并重新写入
        db_handler.execute("DELETE FROM scheduler_jobs")

        for job in scheduler.get_jobs():
            cron_expression = _extract_cron_expression(job.trigger)
            job_id = job.id
            sort_order = existing_orders.get(job_id, 0)

            # 记录保存的CRON表达式
            logger.info(f"保存定时任务到数据库: {job.name}, CRON表达式: {cron_expression}")

            db_handler.execute(
                """INSERT INTO scheduler_jobs (id, name, project_name, module_name, case_index, cron_expression, sort_order)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (
                    job_id,
                    job.name,
                    job.args[0] if len(job.args) > 0 else '',
                    job.args[1] if len(job.args) > 1 else '',
                    job.args[2] if len(job.args) > 2 else 0,
                    cron_expression,
                    sort_order
                )
            )

        logger.info("定时任务已保存到数据库")
    except Exception as e:
        logger.error(f"保存定时任务配置失败: {e}")


def _update_scheduler_jobs_project_name(old_name: str, new_name: str) -> None:
    """
    项目重命名时，同步更新相关定时任务的 project_name

    Args:
        old_name: 旧项目名称
        new_name: 新项目名称
    """
    try:
        # 先收集需要更新的任务信息，避免遍历时修改
        jobs_to_update = []
        for job in scheduler.get_jobs():
            if job.func == scheduled_test_job and len(job.args) > 0 and job.args[0] == old_name:
                jobs_to_update.append({
                    'id': job.id,
                    'args': list(job.args),
                    'name': job.name,
                    'cron_expr': _extract_cron_expression(job.trigger)
                })
        # 统一更新
        for job_info in jobs_to_update:
            new_args = job_info['args']
            new_args[0] = new_name
            new_job_name = job_info['name'].replace(f"{old_name}/", f"{new_name}/", 1)
            scheduler.remove_job(job_info['id'])
            scheduler.add_job(
                func=scheduled_test_job,
                trigger=CronTrigger.from_crontab(job_info['cron_expr']),
                id=job_info['id'],
                args=new_args,
                name=new_job_name
            )
        if jobs_to_update:
            _save_scheduler_jobs()
            logger.info(f"项目重命名({old_name} -> {new_name})，已同步更新 {len(jobs_to_update)} 个定时任务")
    except Exception as e:
        logger.error(f"更新定时任务项目名称失败: {e}", exc_info=True)


def _update_scheduler_jobs_module_name(project_name: str, old_name: str, new_name: str) -> None:
    """
    模块重命名时，同步更新相关定时任务的 module_name

    Args:
        project_name: 项目名称
        old_name: 旧模块名称
        new_name: 新模块名称
    """
    try:
        # 先收集需要更新的任务信息，避免遍历时修改
        jobs_to_update = []
        for job in scheduler.get_jobs():
            if (job.func == scheduled_test_job and len(job.args) > 1
                    and job.args[0] == project_name and job.args[1] == old_name):
                jobs_to_update.append({
                    'id': job.id,
                    'args': list(job.args),
                    'name': job.name,
                    'cron_expr': _extract_cron_expression(job.trigger)
                })
        # 统一更新
        for job_info in jobs_to_update:
            new_args = job_info['args']
            new_args[1] = new_name
            new_job_name = job_info['name'].replace(f"/{old_name} ", f"/{new_name} ", 1)
            scheduler.remove_job(job_info['id'])
            scheduler.add_job(
                func=scheduled_test_job,
                trigger=CronTrigger.from_crontab(job_info['cron_expr']),
                id=job_info['id'],
                args=new_args,
                name=new_job_name
            )
        if jobs_to_update:
            _save_scheduler_jobs()
            logger.info(f"模块重命名({project_name}/{old_name} -> {project_name}/{new_name})，已同步更新 {len(jobs_to_update)} 个定时任务")
    except Exception as e:
        logger.error(f"更新定时任务模块名称失败: {e}", exc_info=True)


def _update_scheduler_jobs_case_name(project_name: str, module_name: str, api_id: int, old_case_name: str, new_case_name: str) -> None:
    """
    接口重命名时，同步更新相关定时任务的 case_name

    Args:
        project_name: 项目名称
        module_name: 模块名称
        api_id: 接口ID
        old_case_name: 旧接口名称
        new_case_name: 新接口名称
    """
    try:
        # 先收集需要更新的任务信息，避免遍历时修改
        jobs_to_update = []
        for job in scheduler.get_jobs():
            if (job.func == scheduled_test_job and len(job.args) > 2
                    and job.args[0] == project_name and job.args[1] == module_name
                    and job.args[2] == api_id):
                jobs_to_update.append({
                    'id': job.id,
                    'args': list(job.args),
                    'name': job.name,
                    'cron_expr': _extract_cron_expression(job.trigger)
                })
        # 统一更新
        for job_info in jobs_to_update:
            new_job_name = job_info['name'].replace(f" - {old_case_name}", f" - {new_case_name}", 1)
            scheduler.remove_job(job_info['id'])
            scheduler.add_job(
                func=scheduled_test_job,
                trigger=CronTrigger.from_crontab(job_info['cron_expr']),
                id=job_info['id'],
                args=job_info['args'],
                name=new_job_name
            )
        if jobs_to_update:
            _save_scheduler_jobs()
            logger.info(f"接口重命名({project_name}/{module_name} - {old_case_name} -> {new_case_name})，已同步更新 {len(jobs_to_update)} 个定时任务")
    except Exception as e:
        logger.error(f"更新定时任务接口名称失败: {e}", exc_info=True)


def _delete_scheduler_jobs_by_project(project_name: str) -> None:
    """
    删除项目时，清理关联的所有定时任务

    Args:
        project_name: 项目名称
    """
    try:
        # 先收集需要删除的任务ID，避免遍历时修改
        job_ids_to_delete = []
        for job in scheduler.get_jobs():
            if job.func == scheduled_test_job and len(job.args) > 0 and job.args[0] == project_name:
                job_ids_to_delete.append(job.id)
        for job_id in job_ids_to_delete:
            scheduler.remove_job(job_id)
        if job_ids_to_delete:
            _save_scheduler_jobs()
            logger.info(f"删除项目({project_name})，已清理 {len(job_ids_to_delete)} 个关联定时任务")
    except Exception as e:
        logger.error(f"清理项目定时任务失败: {e}", exc_info=True)


def _delete_scheduler_jobs_by_module(project_name: str, module_name: str) -> None:
    """
    删除模块时，清理关联的所有定时任务

    Args:
        project_name: 项目名称
        module_name: 模块名称
    """
    try:
        # 先收集需要删除的任务ID，避免遍历时修改
        job_ids_to_delete = []
        for job in scheduler.get_jobs():
            if (job.func == scheduled_test_job and len(job.args) > 1
                    and job.args[0] == project_name and job.args[1] == module_name):
                job_ids_to_delete.append(job.id)
        for job_id in job_ids_to_delete:
            scheduler.remove_job(job_id)
        if job_ids_to_delete:
            _save_scheduler_jobs()
            logger.info(f"删除模块({project_name}/{module_name})，已清理 {len(job_ids_to_delete)} 个关联定时任务")
    except Exception as e:
        logger.error(f"清理模块定时任务失败: {e}", exc_info=True)


def _extract_cron_expression(trigger):
    """从CronTrigger中提取标准cron表达式（分 时 日 月 周）"""
    import re

    # 方法1：从trigger.fields提取
    if hasattr(trigger, 'fields'):
        cron_parts = []
        for field_name in ['minute', 'hour', 'day', 'month', 'day_of_week']:
            field = getattr(trigger.fields, field_name, None)
            if field is not None:
                cron_parts.append(str(field))
        if cron_parts and len(cron_parts) == 5:
            return ' '.join(cron_parts)

    # 方法2：从trigger的str表示解析
    trigger_str = str(trigger)
    # trigger格式如: "cron[minute='*/5', hour='*']"
    cron_match = re.search(r"cron\[(.+)\]", trigger_str)
    if cron_match:
        parts = {}
        for m in re.finditer(r"(\w+)='([^']*)'", cron_match.group(1)):
            parts[m.group(1)] = m.group(2)
        cron_expr = ' '.join([
            parts.get('minute', '*'),
            parts.get('hour', '*'),
            parts.get('day', '*'),
            parts.get('month', '*'),
            parts.get('day_of_week', '*')
        ])
        if cron_expr != '* * * * *' or parts:  # 确保解析到了有效数据
            return cron_expr

    return ''


def _load_scheduler_jobs():
    """从数据库恢复定时任务"""
    logger.info("=" * 50)
    logger.info("开始从数据库恢复定时任务")
    logger.info("=" * 50)

    if not db_handler:
        logger.warning("db_handler 为 None，无法恢复定时任务")
        return

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接已确保")

        jobs_data = db_handler.query("SELECT * FROM scheduler_jobs ORDER BY sort_order, id")
        logger.info(f"从数据库查询到 {len(jobs_data) if jobs_data else 0} 条定时任务记录")

        if not jobs_data:
            logger.info("数据库中没有定时任务数据")
            return

        restored_count = 0
        for job_info in jobs_data:
            try:
                logger.info(f"处理定时任务记录: {job_info}")
                cron_expr = (job_info.get('cron_expression') or '').strip()
                if not cron_expr:
                    logger.warning(f"跳过无效定时任务({job_info.get('name', '未知')}): cron表达式为空")
                    continue

                # 记录加载的CRON表达式
                logger.info(f"从数据库加载定时任务: {job_info.get('name', '未知')}, CRON表达式: {cron_expr}")
                logger.info(f"任务参数: project_name={job_info.get('project_name')}, module_name={job_info.get('module_name')}, case_index={job_info.get('case_index')}")

                scheduler.add_job(
                    func=scheduled_test_job,
                    trigger=CronTrigger.from_crontab(cron_expr),
                    id=job_info['id'],
                    args=[job_info['project_name'], job_info['module_name'], job_info.get('case_index', 0)],
                    name=job_info['name']
                )
                # 触发一次next_run_time的计算
                job = scheduler.get_job(job_info['id'])
                if job and job.next_run_time:
                    logger.info(f"定时任务 {job_info.get('name', '未知')} 下次执行时间: {job.next_run_time.strftime('%Y-%m-%d %H:%M:%S')}")
                else:
                    logger.warning(f"定时任务 {job_info.get('name', '未知')} next_run_time 为空")
                restored_count += 1
            except Exception as e:
                logger.warning(f"恢复定时任务失败({job_info.get('name', '未知')}): {e}", exc_info=True)

        logger.info(f"已恢复 {restored_count} 个定时任务")
        logger.info("=" * 50)
    except Exception as e:
        logger.error(f"加载定时任务配置失败: {e}", exc_info=True)


# 定时任务调度器
scheduler = BackgroundScheduler()
scheduler.start()


def _check_db_available() -> bool:
    """
    检查数据库是否可用（统一连接检查，避免重复 ping 导致 cursor 失效）

    Returns:
        bool: 数据库是否可用
    """
    if not db_handler:
        logger.debug("_check_db_available: db_handler 为 None，数据库不可用")
        return False
    try:
        db_handler._ensure_connection()
        available = db_handler.connection is not None
        logger.debug(f"_check_db_available: 数据库可用={available}, connection={db_handler.connection is not None}, cursor={db_handler.cursor is not None}")
        return available
    except Exception as e:
        logger.warning(f"数据库连接检查失败: {e}")
        return False


def get_statistics_data() -> List[Dict[str, Any]]:
    """
    获取统计数据（仅从数据库读取）

    Returns:
        统计数据列表
    """
    try:
        if not db_handler:
            logger.error("db_handler 为 None，无法读取统计数据")
            return []

        # 确保连接可用
        db_handler._ensure_connection()

        # 从数据库读取统计数据
        statistics = db_handler.query(
            """SELECT id, method, url, status_code, response_time,
               assertion_passed, assertion_count, assertion_passed_count,
               source, project, module, case_name, request_headers,
               request_body, response_headers, response_body,
               timestamp, error, created_at
               FROM test_statistics
               ORDER BY id DESC"""
        )

        if not statistics:
            return []

        # 批量获取所有断言结果（消除N+1查询问题）
        stat_ids = [int(stat['id']) for stat in statistics if stat.get('id')]
        if stat_ids:
            placeholders = ','.join(['%s'] * len(stat_ids))
            all_assertions = db_handler.query(
                f"SELECT * FROM assertion_results WHERE statistic_id IN ({placeholders})",
                tuple(stat_ids)
            )
            # 按 statistic_id 分组
            assertion_map = {}
            for ar in all_assertions:
                sid = ar.get('statistic_id')
                if sid not in assertion_map:
                    assertion_map[sid] = []
                assertion_map[sid].append(ar)
        else:
            assertion_map = {}

        # 为每条记录添加断言结果，并做字段映射确保前端兼容
        for stat in statistics:
            # 映射 timestamp：优先用 timestamp 字段，否则用 created_at
            if 'timestamp' not in stat or not stat['timestamp']:
                stat['timestamp'] = str(stat.get('created_at', '')) if stat.get('created_at') else ''

            # 确保 error 字段存在
            if 'error' not in stat:
                stat['error'] = ''

            # 从预加载的断言结果映射中获取
            statistic_id = int(stat['id']) if stat['id'] else 0
            assertions = assertion_map.get(statistic_id, [])
            # 确保passed字段是布尔类型
            for assertion in assertions:
                if 'passed' in assertion:
                    # 将整数转换为布尔值
                    assertion['passed'] = bool(assertion['passed'])
            stat['assertion_results'] = assertions

        # 确保返回的是列表
        return list(statistics) if isinstance(statistics, tuple) else statistics

    except Exception as e:
        logger.error(f"从数据库读取统计数据失败: {e}")
        return []


def _unwrap_raw(data: Any) -> Any:
    """
    解包 _safe_json_value 包装的 _raw 数据。

    如果数据是字典且仅包含 '_raw' 键，则返回原始纯文本值。
    否则返回原数据。

    Args:
        data: 解析后的 JSON 数据

    Returns:
        解包后的数据
    """
    if isinstance(data, dict) and len(data) == 1 and '_raw' in data:
        return data['_raw']
    return data


def _safe_json_value(value: Any, ensure_ascii: bool = False) -> str:
    """
    安全地将值序列化为 JSON 字符串。

    对于字典/列表：直接 json.dumps
    对于字符串：尝试解析为 JSON，成功则原样序列化；失败则包装为 {"_raw": 原始文本}
    对于 None/空值：返回 '{}'

    Args:
        value: 要序列化的值
        ensure_ascii: 是否转义非 ASCII 字符

    Returns:
        str: 合法的 JSON 字符串
    """
    if value is None or value == '':
        return '{}'
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=ensure_ascii)
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            # 解析成功，确保顶层是对象或数组
            if isinstance(parsed, (dict, list)):
                return json.dumps(parsed, ensure_ascii=ensure_ascii)
            # 顶层是字符串/数字/布尔值，包装为对象
            return json.dumps({'_raw': value}, ensure_ascii=ensure_ascii)
        except (json.JSONDecodeError, ValueError):
            # 纯文本，包装为对象保留原始内容
            return json.dumps({'_raw': value}, ensure_ascii=ensure_ascii)
    # 其他类型（数字、布尔等），包装为对象
    return json.dumps({'_raw': str(value)}, ensure_ascii=ensure_ascii)


def _insert_statistic_record_to_db(record: Dict[str, Any]) -> bool:
    """
    将单条统计记录插入数据库（内部辅助函数）

    Args:
        record: 统计记录

    Returns:
        bool: 插入是否成功
    """
    if not db_handler:
        logger.error("[DB_INSERT] db_handler 为 None，无法插入")
        return False

    # 确保数据库连接可用
    try:
        db_handler._ensure_connection()
    except Exception as e1:
        logger.warning(f"[DB_INSERT] _ensure_connection 失败: {e1}，尝试 connect")
        try:
            db_handler.connect()
        except Exception as e2:
            logger.error(f"[DB_INSERT] connect 也失败: {e2}")
            return False

    if not db_handler.connection:
        logger.error("[DB_INSERT] db_handler.connection 仍为 None，无法插入")
        return False

    # 插入测试统计记录（包含 timestamp 和 error 字段，确保数据完整写入数据库）
    db_handler.execute(
        """INSERT INTO test_statistics
        (method, url, status_code, response_time, assertion_passed, assertion_count,
        assertion_passed_count, source, project, module, case_name, request_headers,
        request_body, response_headers, response_body, timestamp, error)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (
            record.get('method', '') or 'GET',
            record.get('url', '') or 'http://localhost',
            record.get('status_code') or 0,
            record.get('response_time') or 0.0,
            record.get('assertion_passed') if record.get('assertion_passed') is not None else False,
            record.get('assertion_count') or 0,
            record.get('assertion_passed_count') or 0,
            record.get('source', '') or '手动',
            record.get('project', '') or '默认项目',
            record.get('module', '') or '默认模块',
            record.get('case_name', '') or '默认用例',
            json.dumps(record.get('request_headers', {}) or {}, ensure_ascii=False),
            _safe_json_value(record.get('request_body'), ensure_ascii=False),
            json.dumps(record.get('response_headers', {}) or {}, ensure_ascii=False),
            _safe_json_value(record.get('response_body'), ensure_ascii=False),
            record.get('timestamp', '') or datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            record.get('error', '') or ''
        )
    )

    # 获取刚插入的统计ID
    result = db_handler.query(
        "SELECT id FROM test_statistics ORDER BY id DESC LIMIT 1"
    )
    if result and result[0] and result[0]['id']:
        statistic_id = int(result[0]['id'])
    else:
        statistic_id = 0

    # 插入断言结果
    for assertion in record.get('assertion_results', []):
        db_handler.execute(
            """INSERT INTO assertion_results
            (statistic_id, type, field, expected, actual, passed)
            VALUES (%s, %s, %s, %s, %s, %s)""",
            (
                statistic_id,
                assertion.get('type', ''),
                assertion.get('field', ''),
                str(assertion.get('expected', '')),
                str(assertion.get('actual', '')),
                assertion.get('passed', False)
            )
        )
    return True


def save_statistics_data(data: List[Dict[str, Any]]) -> bool:
    """
    保存统计数据（仅操作数据库）

    当 data 为空列表时清空数据库表（用户主动清空统计数据的场景）。
    否则逐条插入新数据（增量模式）。

    Args:
        data: 统计数据列表

    Returns:
        bool: 保存是否成功
    """
    try:
        if not db_handler:
            logger.error("db_handler 为 None，无法保存统计数据")
            return False

        # 确保连接可用
        db_handler._ensure_connection()

        if len(data) == 0:
            # 清空数据库表
            db_handler.execute("DELETE FROM assertion_results")
            db_handler.execute("DELETE FROM test_statistics")
            logger.info("统计数据已清空")
        else:
            # 逐条插入新数据（增量模式）
            for record in data:
                _insert_statistic_record_to_db(record)
            logger.info(f"统计数据已追加到数据库，共 {len(data)} 条记录")

        return True

    except Exception as e:
        logger.error(f"保存统计数据时发生异常: {str(e)}")
        return False
            

def add_statistics_record(record: Dict[str, Any]) -> bool:
    """
    添加一条统计记录，带有保存状态检查

    Args:
        record: 统计记录
        
    Returns:
        bool: 添加并保存是否成功
    """
    try:
        record['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 直接尝试数据库插入
        if db_handler:
            try:
                insert_result = _insert_statistic_record_to_db(record)
                if insert_result:
                    logger.info(f"统计记录已插入数据库, source={record.get('source', '未知')}")
                    return True
                else:
                    logger.warning("_insert_statistic_record_to_db 返回 False，尝试重新连接")
            except Exception as db_err:
                logger.warning(f"数据库插入失败: {db_err}，尝试重新连接")

            # 重试一次
            try:
                db_handler.connect()
                retry_result = _insert_statistic_record_to_db(record)
                if retry_result:
                    logger.info(f"重新连接后统计记录已插入数据库, source={record.get('source', '未知')}")
                    return True
                else:
                    logger.error("重新连接后插入仍返回 False")
                    return False
            except Exception as retry_err:
                logger.error(f"重新连接数据库后插入仍失败: {retry_err}")
                return False

        logger.error("db_handler 为 None，无法插入统计记录")
        return False

    except Exception as e:
        logger.error(f"添加统计记录时发生异常: {str(e)}")
        return False


def sync_file_records_to_db() -> int:
    """
    已废弃：不再使用文件同步，数据全部在数据库中操作。
    保留函数签名以兼容调用点。

    Returns:
        int: 0
    """
    return 0


def get_test_data() -> Dict[str, Any]:
    """
    获取测试数据（从数据库读取，兼容旧格式）

    Returns:
        测试数据字典，格式为 {project_name: [api_list]}
    """
    try:
        if not db_handler:
            return {}

        db_handler._ensure_connection()

        # 从数据库读取所有项目、模块、API，组装成兼容旧格式的字典
        result = {}
        projects = db_handler.query("SELECT id, name FROM projects")
        for proj in projects:
            proj_name = proj['name']
            proj_id = proj['id']

            modules = db_handler.query("SELECT id, name FROM modules WHERE project_id = %s", (proj_id,))
            api_list = []
            for mod in modules:
                mod_id = mod['id']
                apis = db_handler.query("SELECT * FROM apis WHERE module_id = %s ORDER BY id", (mod_id,))
                for api in apis:
                    api_list.append({
                        'case_name': api.get('case_name', ''),
                        'url': api.get('url', ''),
                        'method': api.get('method', 'GET'),
                        'headers': api.get('headers') if isinstance(api.get('headers'), dict) else {},
                        'data': api.get('data') if isinstance(api.get('data'), (dict, str)) else {},
                        'expected': api.get('expected') if isinstance(api.get('expected'), dict) else {},
                        'extractions': api.get('extractions') if isinstance(api.get('extractions'), dict) else {},
                        '_module_name': mod['name'],  # 附加模块名信息
                    })

            if api_list:
                result[proj_name] = api_list

        return result

    except Exception as e:
        logger.error(f"从数据库读取测试数据失败: {e}")
        return {}


def save_test_data(data: Dict[str, Any]) -> bool:
    """
    保存测试数据（已废弃，数据通过各路由直接操作数据库）
    保留函数签名以兼容调用点。

    Args:
        data: 测试数据字典

    Returns:
        是否保存成功
    """
    logger.warning("save_test_data() 已废弃，数据应通过各路由直接操作数据库")
    return True


def get_projects() -> Dict[str, Any]:
    # Use fixed function to always get latest data
    return get_projects_fixed()


def _get_project_current_env(project_id: int) -> str:
    """获取项目当前激活的环境名称"""
    try:
        envs = db_handler.query("SELECT name FROM environments WHERE project_id = %s ORDER BY id", (project_id,))
        if envs:
            return envs[0]['name']
    except Exception:
        pass
    return ''


def save_projects(data: Dict[str, Any]) -> bool:
    """
    保存项目数据（已废弃，数据通过各路由直接操作数据库）
    保留函数签名以兼容调用点。

    Args:
        data: 项目数据字典

    Returns:
        是否保存成功
    """
    logger.warning("save_projects() 已废弃，数据应通过各路由直接操作数据库")
    return True


def execute_api_test(api_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    执行API测试

    Args:
        api_data: API测试数据

    Returns:
        测试结果
    """
    try:
        if not isinstance(api_data, dict):
            logger.error(f"api_data类型错误: type={type(api_data)}")
            return {"error": f"api_data类型错误: 期望dict, 实际{type(api_data).__name__}", "success": False}

        # 获取项目级别的环境配置
        project_name = api_data.get('project', '')
        base_url = ''
        timeout = 30
        default_headers = {}

        if project_name:
            projects_data_for_env = get_projects()
            if project_name in projects_data_for_env:
                proj = projects_data_for_env[project_name]
                proj_envs = proj.get('envs', {})
                proj_current_env = proj.get('current_env', '')
                # 如果项目有环境配置，使用项目的
                if proj_envs and proj_current_env and proj_current_env in proj_envs:
                    env_cfg = proj_envs[proj_current_env]
                    if isinstance(env_cfg, dict):
                        base_url = env_cfg.get('base_url', '')
                        timeout = env_cfg.get('timeout', 30)
                        default_headers = env_cfg.get('headers', {})

        # 如果项目没有配置环境，base_url 为空，接口URL必须包含完整域名

        # 获取请求数据
        url = api_data.get("url", "")
        method = api_data.get("method", "GET").upper()
        headers = api_data.get("headers", {})
        data = api_data.get("data", {})
        expected = api_data.get("expected", {})

        # 确保数据类型正确（YAML或前端可能传入字符串）
        if isinstance(headers, str):
            try:
                headers = json.loads(headers)
            except (json.JSONDecodeError, TypeError):
                headers = {}
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except (json.JSONDecodeError, TypeError):
                # 纯文本请求体，保留原始字符串
                pass
        if isinstance(expected, str):
            try:
                expected = json.loads(expected)
            except (json.JSONDecodeError, TypeError):
                expected = {}
        if not isinstance(headers, dict):
            headers = {}
        if not isinstance(data, (dict, str)):
            data = {}
        if not isinstance(expected, dict):
            expected = {}

        # 合并默认headers和自定义headers
        merged_headers = {**default_headers, **headers}

        # 处理项目自定义变量替换 {var_name}
        # 始终从数据库读取最新变量，确保拿到前序接口提取的最新值（如token）
        proj_variables = {}
        if project_name and db_handler:
            try:
                db_handler._ensure_connection()
                proj_row = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
                if proj_row:
                    proj_id_for_vars = proj_row[0]['id']
                    vars_rows = db_handler.query("SELECT name, value FROM variables WHERE project_id = %s", (proj_id_for_vars,))
                    proj_variables = {var['name']: var['value'] for var in vars_rows}
                    logger.info(f"[变量替换] 从数据库读取最新项目变量: {proj_variables}")
            except Exception as e:
                logger.error(f"[变量替换] 读取项目变量失败: {e}", exc_info=True)
                # 降级：从项目配置缓存读取
                if project_name in projects_data_for_env:
                    proj_variables = dict(projects_data_for_env[project_name].get('variables', {}))

        if proj_variables:
            import re
            # 变量名模式，匹配 {var_name}
            var_pattern = re.compile(r'\{(\w+)\}')

            def replace_vars(obj, variables, max_rounds=10):
                """递归替换对象中的 {变量名} 占位符，支持嵌套变量（如 Authorization: Bearer {access_token}）"""
                if isinstance(obj, str):
                    for _ in range(max_rounds):
                        matches = var_pattern.findall(obj)
                        if not matches:
                            break
                        changed = False
                        for var_key in matches:
                            if var_key in variables:
                                obj = obj.replace('{' + str(var_key) + '}', str(variables[var_key]))
                                changed = True
                        if not changed:
                            break
                    return obj
                elif isinstance(obj, dict):
                    return {k: replace_vars(v, variables, max_rounds) for k, v in obj.items()}
                elif isinstance(obj, list):
                    return [replace_vars(item, variables, max_rounds) for item in obj]
                return obj
            url = replace_vars(url, proj_variables)
            headers = replace_vars(headers, proj_variables)
            data = replace_vars(data, proj_variables)
            merged_headers = {**default_headers, **headers}

        # 处理数据中的动态变量
        if "${timestamp}" in str(data) or "${timestamp}" in str(headers):
            timestamp = int(time.time())
            if isinstance(data, dict):
                for key, value in data.items():
                    if isinstance(value, str) and "${timestamp}" in value:
                        data[key] = value.replace("${timestamp}", str(timestamp))
            elif isinstance(data, str) and "${timestamp}" in data:
                data = data.replace("${timestamp}", str(timestamp))
            for key, value in headers.items():
                if isinstance(value, str) and "${timestamp}" in value:
                    headers[key] = value.replace("${timestamp}", str(timestamp))
                    merged_headers[key] = headers[key]

        # 拼接完整URL
        full_url = url
        if not url.startswith("http"):
            full_url = f"{base_url}{url}"

        import time as time_module

        # 使用独立会话，避免不同任务之间的token相互影响
        # 从api_data中获取任务ID，如果没有则使用默认值
        task_id = api_data.get('task_id', 'default')
        task_session = session_manager.get_session(task_id)

        # 移除连接预热机制以提高响应速度
        try:
            from urllib3.util.url import parse_url as _parse_url
            _parsed = _parse_url(full_url)
            _warmup_url = f"{_parsed.scheme}://{_parsed.auth or ''}{_parsed.host}{':' + str(_parsed.port) if _parsed.port else ''}/"
            task_session.request("HEAD", _warmup_url, timeout=5)
        except Exception:
            pass  # 预热失败不影响后续请求

        # 记录发送请求的时间（预热之后，确保连接已建立）
        start_time = time_module.perf_counter()

        # 根据data类型选择发送方式：字典用json，字符串用data（纯文本请求体）
        if isinstance(data, str):
            request_kwargs = {'data': data, 'headers': merged_headers, 'timeout': timeout}
        else:
            request_kwargs = {'json': data, 'headers': merged_headers, 'timeout': timeout}

        if method == "GET":
            raw_response = task_session.request("GET", full_url, params=data, headers=merged_headers, timeout=timeout)
        elif method == "POST":
            raw_response = task_session.request("POST", full_url, **request_kwargs)
        elif method == "PUT":
            raw_response = task_session.request("PUT", full_url, **request_kwargs)
        elif method == "DELETE":
            raw_response = task_session.request("DELETE", full_url, headers=merged_headers, timeout=timeout)
        else:
            raise ValueError(f"不支持的请求方法: {method}")

        # 响应时间 = 从发送请求到接收完响应体的完整耗时（不含断言处理时间、不含连接预热时间）
        end_time = time_module.perf_counter()
        response_time = round((end_time - start_time) * 1000, 2)  # 毫秒

        # 解析响应数据
        try:
            if raw_response.headers.get("content-type", "").startswith("application/json"):
                response_data = raw_response.json()
            else:
                response_data = raw_response.text
        except Exception:
            response_data = raw_response.text

        response = {
            "status_code": raw_response.status_code,
            "headers": dict(raw_response.headers),
            "data": response_data
        }

        # 执行断言验证（不计入响应时间）
        assertion_results = []
        assertion_passed = True
        
        # 验证状态码
        if "status_code" in expected:
            expected_status_code = expected["status_code"]
            actual_status_code = response.get("status_code")
            # 确保类型一致，都转换为整数进行比较
            try:
                expected_status_code_int = int(expected_status_code)
                actual_status_code_int = int(actual_status_code)
                if expected_status_code_int == actual_status_code_int:
                    assertion_results.append({
                        "type": "status_code",
                        "field": "status_code",
                        "expected": expected_status_code_int,
                        "actual": actual_status_code_int,
                        "passed": True
                    })
                else:
                    assertion_passed = False
                    assertion_results.append({
                        "type": "status_code",
                        "field": "status_code",
                        "expected": expected_status_code_int,
                        "actual": actual_status_code_int,
                        "passed": False
                    })
            except (ValueError, TypeError) as e:
                logger.warning(f"状态码比较失败: {e}, expected={expected_status_code}, actual={actual_status_code}")
                assertion_passed = False
                assertion_results.append({
                    "type": "status_code",
                    "field": "status_code",
                    "expected": expected_status_code,
                    "actual": actual_status_code,
                    "passed": False
                })
        
        # 验证响应数据
        if "data" in expected:
            expected_data = expected["data"]
            actual_data = response.get("data", {})
            # 确保expected_data和actual_data是字典类型
            if isinstance(expected_data, str):
                try:
                    expected_data = json.loads(expected_data)
                except (json.JSONDecodeError, TypeError):
                    expected_data = {}
            if isinstance(actual_data, str):
                try:
                    actual_data = json.loads(actual_data)
                except (json.JSONDecodeError, TypeError):
                    actual_data = {}
            if not isinstance(expected_data, dict):
                expected_data = {}
            if not isinstance(actual_data, dict):
                actual_data = {}
            
            
            
            for key, expected_value in expected_data.items():
                if key in actual_data:
                    actual_value = actual_data[key]
                    # 确保类型一致，都转换为字符串进行比较
                    expected_str = str(expected_value)
                    actual_str = str(actual_value)
                    if expected_str == actual_str:
                        assertion_results.append({
                            "type": "data",
                            "field": key,
                            "expected": expected_str,
                            "actual": actual_str,
                            "passed": True
                        })
                    else:
                        assertion_passed = False
                        assertion_results.append({
                            "type": "data",
                            "field": key,
                            "expected": expected_str,
                            "actual": actual_str,
                            "passed": False
                        })
                else:
                    assertion_passed = False
                    assertion_results.append({
                        "type": "data",
                        "field": key,
                        "expected": expected_value,
                        "actual": "字段不存在",
                        "passed": False
                    })

        # 执行变量提取
        extraction_results = []
        if "extractions" in api_data:
            extractions = api_data["extractions"]
            if isinstance(extractions, dict):
                # 处理每个提取项
                for var_name, extraction_config in extractions.items():
                    if isinstance(extraction_config, dict) and "path" in extraction_config:
                        path = extraction_config["path"]
                        default_value = extraction_config.get("default", None)

                        # 尝试从响应中提取值
                        extracted_value = None
                        try:
                            # 支持简单路径如 "data.user.id"
                            if path:
                                parts = path.split(".")
                                current = response.get("data", {})

                                # 确保current是字典类型
                                if isinstance(current, str):
                                    try:
                                        current = json.loads(current)
                                    except (json.JSONDecodeError, TypeError):
                                        current = {}

                                # 遍历路径
                                for part in parts:
                                    if isinstance(current, dict) and part in current:
                                        current = current[part]
                                    else:
                                        current = None
                                        break

                                extracted_value = current if current is not None else default_value
                        except Exception as e:
                            logger.warning(f"提取变量 {var_name} 时出错: {e}")
                            extracted_value = default_value

                        # 记录提取结果
                        if extracted_value is not None:
                            extraction_results.append({
                                "var_name": var_name,
                                "path": path,
                                "value": extracted_value,
                                "success": True
                            })
                            # 立即保存到数据库，确保后续接口能拿到最新变量
                            if project_name and db_handler:
                                try:
                                    db_handler._ensure_connection()
                                    proj_row = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
                                    if proj_row:
                                        proj_id_for_save = proj_row[0]['id']
                                        existing_var = db_handler.query(
                                            "SELECT id FROM variables WHERE project_id = %s AND name = %s",
                                            (proj_id_for_save, var_name)
                                        )
                                        if existing_var:
                                            db_handler.execute(
                                                "UPDATE variables SET value = %s WHERE project_id = %s AND name = %s",
                                                (str(extracted_value), proj_id_for_save, var_name)
                                            )
                                        else:
                                            db_handler.execute(
                                                "INSERT INTO variables (project_id, name, value) VALUES (%s, %s, %s)",
                                                (proj_id_for_save, var_name, str(extracted_value))
                                            )
                                        logger.info(f"[变量提取] 立即保存到数据库: {var_name} = {extracted_value}")
                                        # 清除缓存，确保后续接口读取到最新值
                                        clear_project_cache()
                                except Exception as e:
                                    logger.error(f"[变量提取] 保存变量到数据库失败: {e}", exc_info=True)
                        else:
                            extraction_results.append({
                                "var_name": var_name,
                                "path": path,
                                "value": None,
                                "success": False
                            })

        # 构建显示用的完整URL（对于GET请求，将params拼接到URL中显示）
        display_url = full_url
        if method == "GET" and data:
            query_string = "&".join([f"{k}={v}" for k, v in data.items()])
            display_url = f"{display_url}?{query_string}" if "?" not in display_url else f"{display_url}&{query_string}"

        # 记录统计数据
        try:
            stat_record = {
                "method": method,
                "url": display_url,
                "status_code": response.get("status_code"),
                "response_time": response_time,
                "assertion_passed": assertion_passed,
                "assertion_count": len(assertion_results),
                "assertion_passed_count": sum(1 for a in assertion_results if a.get("passed")),
                "assertion_results": assertion_results,
                "source": api_data.get("source", "手动调试"),
                "project": api_data.get("project", ""),
                "module": api_data.get("module", ""),
                "case_name": api_data.get("case_name", ""),
                "request_headers": merged_headers,
                "request_body": data if data else None,
                "response_headers": response.get("headers", {}),
                "response_body": response.get("data"),
            }
            logger.info(f"[execute_api_test] 准备记录统计数据, source={stat_record.get('source')}, url={display_url}")
            # 检查统计数据是否成功记录
            stat_saved = add_statistics_record(stat_record)
            logger.info(f"[execute_api_test] 统计数据记录结果: saved={stat_saved}")
            if not stat_saved:
                logger.error("记录统计数据失败：无法保存到数据库")
                # 可以在这里添加额外的错误处理逻辑，如重试机制
        except Exception as stat_err:
            logger.error(f"记录统计数据时发生异常: {stat_err}")

        # 返回测试结果，包含更新后的项目变量（供批量执行时传递给下一个接口）
        return {
            "request_method": method,
            "request_url": display_url,
            "request_headers": merged_headers,
            "request_body": data if data else None,
            "status_code": response.get("status_code"),
            "headers": response.get("headers", {}),
            "data": response.get("data"),
            "response_time": response_time,
            "assertion_results": assertion_results,
            "assertion_passed": assertion_passed,
            "extraction_results": extraction_results,
            "success": True
        }
    except Exception as e:
        import traceback
        tb_str = traceback.format_exc()
        logger.error(f"执行API测试失败: {e}\n{tb_str}")
        # 将traceback写入调试文件
        try:
            debug_path = os.path.join(BASE_DIR, 'debug_error.log')
            with open(debug_path, 'w', encoding='utf-8') as f:
                f.write(f"=== execute_api_test 异常 ===\n")
                f.write(f"api_data type: {type(api_data)}\n")
                f.write(f"api_data: {api_data}\n")
                f.write(f"error: {e}\n")
                f.write(f"traceback:\n{tb_str}\n")
        except Exception:
            pass
        # 先返回错误结果，再尝试记录统计
        error_result = {
            "error": str(e),
            "traceback": tb_str,
            "success": False
        }
        # 记录失败的统计数据
        try:
            if isinstance(api_data, dict):
                stat_record = {
                    "method": api_data.get("method", "GET").upper(),
                    "url": api_data.get("url", ""),
                    "status_code": None,
                    "response_time": None,
                    "assertion_passed": False,
                    "assertion_count": 0,
                    "assertion_passed_count": 0,
                    "source": api_data.get("source", "手动调试"),
                    "project": api_data.get("project", ""),
                    "module": api_data.get("module", ""),
                    "case_name": api_data.get("case_name", ""),
                    "error": str(e),
                }
                add_statistics_record(stat_record)
        except Exception:
            pass
        return error_result


def scheduled_test_job(project_name: str, module_name: str, api_id: int) -> None:
    """
    定时测试任务
    执行指定模块中所有需要执行的接口（按顺序执行）

    Args:
        project_name: 项目名称
        module_name: 模块名称
        api_id: 触发任务的接口ID（数据库ID），用于标识是哪个job触发的
    """
    logger.info(f"[定时测试] 开始执行定时任务: {project_name}/{module_name}[{api_id}]")

    # 使用锁防止同一模块的多个定时任务并发执行
    _scheduler_module_lock_key = f"{project_name}__{module_name}"
    if not hasattr(scheduled_test_job, '_module_locks'):
        scheduled_test_job._module_locks = {}
    if _scheduler_module_lock_key not in scheduled_test_job._module_locks:
        import threading
        scheduled_test_job._module_locks[_scheduler_module_lock_key] = threading.Lock()

    lock = scheduled_test_job._module_locks[_scheduler_module_lock_key]
    if not lock.acquire(blocking=False):
        logger.info(f"[定时测试] 同模块任务正在执行，跳过: {project_name}/{module_name}[{api_id}]")
        return

    try:
        # 获取项目级别的环境配置 - 直接从数据库获取最新数据，不使用缓存
        projects_data_for_env = _get_projects_directly()
        proj_env_name = ''
        if project_name in projects_data_for_env:
            proj_env_name = projects_data_for_env[project_name].get('current_env', '')

        logger.info(f"[定时测试] 项目环境: {proj_env_name}")

        if project_name and module_name and api_id and db_handler:
            try:
                db_handler._ensure_connection()
                logger.info(f"[定时测试] 查询项目: {project_name}")
                proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
                if proj:
                    proj_id = proj[0]['id']
                    logger.info(f"[定时测试] 查询模块: {module_name}")
                    mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
                    if mod:
                        mod_id = mod[0]['id']

                        # 获取当前时间，判断哪些接口需要执行
                        current_time = datetime.now().astimezone()
                        logger.info(f"[定时测试] 当前时间: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")

                        # 获取该模块下所有设置了定时任务的接口
                        # 直接调用scheduler_list函数获取数据，确保使用相同的next_run_time
                        with app.test_request_context():
                            response = scheduler_list()
                            jobs_data = response.get_json()

                        module_jobs = []
                        for job_data in jobs_data:
                            if (job_data['project_name'] == project_name and
                                job_data['module_name'] == module_name):
                                # 解析next_run_time字符串为datetime对象
                                next_run_time = None
                                if job_data.get('next_run_time'):
                                    try:
                                        next_run_time = datetime.strptime(job_data['next_run_time'], '%Y-%m-%d %H:%M:%S')
                                        # 添加时区信息，使其与current_time兼容
                                        next_run_time = next_run_time.replace(tzinfo=current_time.tzinfo)
                                    except Exception as e:
                                        logger.error(f"解析next_run_time失败: {e}")

                                module_jobs.append({
                                    'api_id': job_data['api_id'],
                                    'job_id': job_data['id'],
                                    'next_run_time': next_run_time,
                                    'sort_order': job_data.get('sort_order', 0)
                                })



                        if not module_jobs:
                            logger.info(f"[定时测试] 模块 {module_name} 下没有定时任务")
                            return

                        # 按sort_order排序，确保按顺序执行
                        module_jobs.sort(key=lambda x: x['sort_order'])
                        logger.info(f"[定时测试] 模块 {module_name} 下有 {len(module_jobs)} 个定时任务需要执行")

                        # 检查哪些任务应该在当前时间执行（允许1分钟的误差）
                        jobs_to_execute = []
                        for job_info in module_jobs:
                            if job_info['next_run_time']:
                                logger.info(f"[定时测试] 检查接口 {job_info['api_id']}: next_run_time={job_info['next_run_time']}, type={type(job_info['next_run_time'])}")
                                logger.info(f"[定时测试] 检查接口 {job_info['api_id']}: current_time={current_time}, type={type(current_time)}")
                                try:
                                    time_diff = abs((current_time - job_info['next_run_time']).total_seconds())
                                    logger.info(f"[定时测试] 接口 {job_info['api_id']} 时间差: {time_diff}秒")
                                    if time_diff <= 60:  # 允许1分钟的误差
                                        jobs_to_execute.append(job_info['api_id'])
                                        logger.info(f"[定时测试] 接口 {job_info['api_id']} 需要执行 (下次执行时间: {job_info['next_run_time'].strftime('%Y-%m-%d %H:%M:%S')})")
                                except Exception as e:
                                    logger.error(f"[定时测试] 计算时间差失败: {e}, current_time={current_time}, next_run_time={job_info['next_run_time']}")

                        if not jobs_to_execute:
                            logger.info(f"[定时测试] 当前时间没有需要执行的接口")
                            return

                        # 按顺序执行所有需要执行的接口
                        for api_id in jobs_to_execute:
                            logger.info(f"[定时测试] 查询接口ID: {api_id}")
                            api = db_handler.query("SELECT * FROM apis WHERE id = %s", (api_id,))
                            if not api:
                                logger.warning(f"定时测试失败: 接口不存在 - {api_id}")
                                continue
                            api_row = api[0]

                            # 执行指定的接口
                            logger.info(f"[定时测试] 执行指定接口: {api_row.get('case_name', '')} (id={api_row['id']})")
                            api_data = {
                                'case_name': api_row.get('case_name', ''),
                                'url': api_row.get('url', ''),
                                'method': api_row.get('method', 'GET'),
                                'headers': json.loads(api_row.get('headers', '{}')),
                                'data': _unwrap_raw(json.loads(api_row.get('data', '{}'))),
                                'expected': json.loads(api_row.get('expected', '{}')),
                                'extractions': json.loads(api_row.get('extractions', '{}')),
                                'project': project_name,
                                'module': module_name,
                                'source': '定时',
                                'task_id': f"{project_name}_{module_name}_{api_row['id']}_{int(time.time())}"
                            }
                            try:
                                result = execute_api_test(api_data)
                                logger.info(f"[定时测试] 执行接口 {api_row.get('case_name', '')} (id={api_row['id']}) 完成, 成功={result.get('success')}")

                                # 检查断言是否失败，如果失败且项目配置了邮箱报警，则发送报警邮件
                                if not result.get('assertion_passed', True):
                                    logger.warning(f"[定时测试] 接口 {api_row.get('case_name', '')} 断言失败")
                                    # 查询项目的邮箱报警配置
                                    proj_alert = db_handler.query(
                                        "SELECT alert_enabled, alert_email FROM projects WHERE name = %s",
                                        (project_name,)
                                    )
                                    if proj_alert and proj_alert[0].get('alert_enabled') == 1 and proj_alert[0].get('alert_email'):
                                        alert_email = proj_alert[0].get('alert_email')
                                        logger.info(f"[定时测试] 发送报警邮件到: {alert_email}")
                                        # 发送报警邮件
                                        try:
                                            email_handler.send_alert_email(
                                                to_email=alert_email,
                                                project_name=project_name,
                                                module_name=module_name,
                                                case_name=api_row.get('case_name', ''),
                                                result=result
                                            )
                                            logger.info(f"[定时测试] 报警邮件发送成功")
                                        except Exception as e:
                                            logger.error(f"[定时测试] 报警邮件发送失败: {e}", exc_info=True)
                                    else:
                                        logger.info(f"[定时测试] 项目未配置邮箱报警，跳过发送邮件")
                            except Exception as e:
                                logger.warning(f"[定时测试] 执行接口 {api_row.get('case_name', '')} (id={api_row['id']}) 失败: {e}")
                    else:
                        logger.warning(f"[定时测试] 模块不存在: {module_name}")
                else:
                    logger.warning(f"[定时测试] 项目不存在: {project_name}")
            except Exception as e:
                logger.error(f"定时测试执行失败: {e}", exc_info=True)
        else:
            logger.error(f"定时测试失败: 参数不完整 - project_name={project_name}, module_name={module_name}, api_id={api_id}, db_handler={db_handler is not None}")
    finally:
        lock.release()
        logger.info(f"[定时测试] 定时任务执行完成: {project_name}/{module_name}")


# 启动时恢复持久化的定时任务（必须在 scheduled_test_job 函数定义之后）
_load_scheduler_jobs()


# 设备检测函数
def is_mobile_device(user_agent):
    """检测是否为移动设备"""
    if not user_agent:
        return False
    mobile_keywords = ['Mobile', 'Android', 'iPhone', 'iPad', 'Windows Phone', 'webOS', 'BlackBerry']
    return any(keyword in user_agent for keyword in mobile_keywords)

# 设备检测和页面渲染函数
def _render_index_page(user_agent, test_data=None, projects_data=None):
    """根据设备类型渲染页面"""
    if is_mobile_device(user_agent):
        return render_template('mobile.html', projects_data=projects_data)
    else:
        return render_template('index.html', test_data=test_data, projects_data=projects_data)

# 路由：首页
@app.route('/')
def index():
    """首页 - 根据设备类型返回不同页面"""
    user_agent = request.headers.get('User-Agent', '')
    
    # 只在需要时才获取数据
    if is_mobile_device(user_agent):
        projects_data = get_projects()
        return _render_index_page(user_agent, projects_data=projects_data)
    else:
        test_data = get_test_data()
        projects_data = get_projects()
        return _render_index_page(user_agent, test_data, projects_data)

# 路由：手机端页面
@app.route('/mobile')
@cache_result('mobile_projects', projects_cache, timeout=60)  # 缓存1分钟
def mobile():
    """手机端页面"""
    projects_data = get_projects()
    return render_template('mobile.html', projects_data=projects_data)


# 路由：顶部统计数据
@app.route('/api/top_stats')
@cache_result('top_stats', stats_cache, timeout=15)  # 缓存15秒
def api_top_stats():
    """获取顶部统计数据（项目数、模块数、接口数、定时任务数）"""

    projects_data = get_projects()
    
    # 项目总数
    project_count = len(projects_data)
    
    # 模块总数和接口总数
    module_count = 0
    api_count = 0
    for project_data in projects_data.values():
        if isinstance(project_data, dict) and 'modules' in project_data:
            modules = project_data['modules']
            module_count += len(modules)
            for module_data in modules.values():
                if isinstance(module_data, dict) and 'apis' in module_data:
                    api_count += len(module_data['apis'])
    
    # 定时任务总数
    scheduler_count = len(scheduler.get_jobs())
    
    return jsonify({
        'project_count': project_count,
        'module_count': module_count,
        'api_count': api_count,
        'scheduler_count': scheduler_count
    })


# 路由：API列表
@app.route('/api/list')
def api_list():
    """API列表"""
    test_data = get_test_data()
    return jsonify(test_data)


# 路由：添加项目
@app.route('/projects/add', methods=['POST'])
def project_add():
    """添加项目"""
    project_name = request.form.get('project_name', '').strip()
    project_desc = request.form.get('project_desc', '').strip()
    alert_enabled = request.form.get('alert_enabled', '0').strip()
    alert_email = request.form.get('alert_email', '').strip()

    logger.info(f"========== 开始添加项目 ==========")
    logger.info(f"项目名称: {project_name}, 描述: {project_desc}, 报警启用: {alert_enabled}, 报警邮箱: {alert_email}")

    if not project_name:
        return jsonify({'success': False, 'error': '项目名称不能为空'})

    if not db_handler:
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 检查项目是否已存在
        existing = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if existing:
            logger.warning(f"项目已存在: {project_name}")
            return jsonify({'success': False, 'error': '项目已存在'})

        # 添加项目
        db_handler.execute(
            "INSERT INTO projects (name, description, alert_enabled, alert_email) VALUES (%s, %s, %s, %s)",
            (project_name, project_desc, alert_enabled, alert_email)
        )
        logger.info("项目添加成功")

        return jsonify({'success': True, 'message': '项目添加成功'})
    except Exception as e:
        logger.error(f"添加项目失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'添加项目失败: {str(e)}'})


# 路由：删除项目
@app.route('/projects/delete/<project_name>', methods=['POST'])
def project_delete(project_name):
    """删除项目"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    logger.info(f"========== 开始删除项目 ==========")
    logger.info(f"项目名称: {project_name}")

    if not db_handler:
        logger.error("数据库未连接")
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 删除项目下的所有模块的接口
        modules = db_handler.query("SELECT id FROM modules WHERE project_id = %s", (proj_id,))
        for mod in modules:
            db_handler.execute("DELETE FROM apis WHERE module_id = %s", (mod['id'],))
        logger.info("项目下的接口删除成功")

        # 删除项目下的所有模块
        db_handler.execute("DELETE FROM modules WHERE project_id = %s", (proj_id,))
        logger.info("项目下的模块删除成功")

        # 删除项目
        db_handler.execute("DELETE FROM projects WHERE id = %s", (proj_id,))
        logger.info("项目删除成功")

        # 删除关联的定时任务
        _delete_scheduler_jobs_by_project(project_name)

        return jsonify({'success': True, 'message': '项目删除成功'})
    except Exception as e:
        logger.error(f"删除项目失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'删除项目失败: {str(e)}'})


# 路由：更新项目
@app.route('/projects/update', methods=['POST'])
def project_update():
    """更新项目"""
    old_name = request.form.get('old_name', '').strip()
    project_name = request.form.get('project_name', '').strip()
    project_desc = request.form.get('project_desc', '').strip()
    alert_enabled = request.form.get('alert_enabled', '0').strip()
    alert_email = request.form.get('alert_email', '').strip()

    logger.info(f"========== 开始更新项目 ==========")
    logger.info(f"旧项目名称: {old_name}, 新项目名称: {project_name}, 描述: {project_desc}, 报警启用: {alert_enabled}, 报警邮箱: {alert_email}")

    if not project_name:
        return jsonify({'success': False, 'error': '项目名称不能为空'})

    if not db_handler:
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (old_name,))
        if not proj:
            logger.warning(f"项目不存在: {old_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 如果修改了项目名称，检查新名称是否已存在
        if old_name != project_name:
            existing = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
            if existing:
                logger.warning(f"项目已存在: {project_name}")
                return jsonify({'success': False, 'error': '项目已存在'})

        # 更新项目
        db_handler.execute(
            "UPDATE projects SET name = %s, description = %s, alert_enabled = %s, alert_email = %s WHERE id = %s",
            (project_name, project_desc, alert_enabled, alert_email, proj_id)
        )
        logger.info("项目更新成功")

        # 如果修改了项目名称，同步更新相关定时任务
        if old_name != project_name:
            _update_scheduler_jobs_project_name(old_name, project_name)

        # 清除项目缓存，确保下次获取最新数据
        clear_project_cache()

        return jsonify({'success': True, 'message': '项目更新成功'})
    except Exception as e:
        logger.error(f"更新项目失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'更新项目失败: {str(e)}'})


# 路由：项目列表
@app.route('/projects/list')
def projects_list():
    """项目列表"""
    logger.info("========== 开始获取项目列表 ==========")

    if not db_handler:
        logger.error("数据库未连接")
        return jsonify({})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询所有项目
        projects = db_handler.query("SELECT * FROM projects")
        logger.info(f"查询到 {len(projects)} 个项目")

        projects_data = {}
        for proj in projects:
            proj_name = proj['name']

            # 查询该项目的所有模块
            modules = db_handler.query("SELECT * FROM modules WHERE project_id = %s", (proj['id'],))
            logger.info(f"项目 {proj_name} 查询到 {len(modules)} 个模块")

            modules_data = {}
            for mod in modules:
                mod_name = mod['name']

                # 查询该模块的所有接口
                apis = db_handler.query("SELECT * FROM apis WHERE module_id = %s", (mod['id'],))
                logger.info(f"模块 {mod_name} 查询到 {len(apis)} 个接口")

                apis_data = []
                for api in apis:
                    apis_data.append({
                        'id': api['id'],  # 添加数据库ID
                        'case_name': api['case_name'],
                        'url': api['url'],
                        'method': api['method'],
                        'headers': json.loads(api['headers']) if api['headers'] else {},
                        'data': _unwrap_raw(json.loads(api['data'])) if api['data'] else {},
                        'expected': json.loads(api['expected']) if api['expected'] else {},
                        'extractions': json.loads(api['extractions']) if api['extractions'] else {}
                    })

                modules_data[mod_name] = {
                    'id': mod['id'],
                    'name': mod['name'],
                    'description': mod['description'],
                    'apis': apis_data
                }

            projects_data[proj_name] = {
                'id': proj['id'],
                'name': proj['name'],
                'description': proj['description'],
                'alert_enabled': proj.get('alert_enabled', 0),
                'alert_email': proj.get('alert_email', ''),
                'modules': modules_data
            }

        logger.info("项目列表获取成功")
        return jsonify(projects_data)
    except Exception as e:
        logger.error(f"获取项目列表失败: {e}", exc_info=True)
        return jsonify({})


@app.route('/projects/<project_name>/modules/delete/<module_name>', methods=['POST'])
def module_delete(project_name, module_name):
    """删除模块"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    module_name = unquote(module_name)
    logger.info(f"========== 开始删除模块 ==========")
    logger.info(f"项目名称: {project_name}, 模块名称: {module_name}")

    if not db_handler:
        logger.error("数据库未连接")
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 查询模块ID
        mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
        if not mod:
            logger.warning(f"模块不存在: {module_name}")
            return jsonify({'success': False, 'error': '模块不存在'})
        mod_id = mod[0]['id']

        # 删除模块下的所有接口
        db_handler.execute("DELETE FROM apis WHERE module_id = %s", (mod_id,))
        logger.info("模块下的接口删除成功")

        # 删除模块
        db_handler.execute("DELETE FROM modules WHERE id = %s", (mod_id,))
        logger.info("模块删除成功")

        # 删除关联的定时任务
        _delete_scheduler_jobs_by_module(project_name, module_name)

        return jsonify({'success': True, 'message': '模块删除成功'})
    except Exception as e:
        logger.error(f"删除模块失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'删除模块失败: {str(e)}'})


@app.route('/projects/<project_name>/modules/add', methods=['POST'])
def module_add(project_name):
    """添加模块"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    module_name = request.form.get('module_name', '').strip()
    module_desc = request.form.get('module_desc', '').strip()
    
    logger.info(f"========== 开始添加模块 ==========")
    logger.info(f"项目名称: {project_name}, 模块名称: {module_name}, 描述: {module_desc}")
    
    if not module_name:
        return jsonify({'success': False, 'error': '模块名称不能为空'})
    
    if not db_handler:
        return jsonify({'success': False, 'error': '数据库未连接'})
    
    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")
        
        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']
        
        # 检查模块是否已存在
        existing = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
        if existing:
            logger.warning(f"模块已存在: {module_name}")
            return jsonify({'success': False, 'error': '模块已存在'})
        
        # 添加模块
        db_handler.execute(
            "INSERT INTO modules (project_id, name, description) VALUES (%s, %s, %s)",
            (proj_id, module_name, module_desc)
        )
        logger.info("模块添加成功")
        
        return jsonify({'success': True, 'message': '模块添加成功'})
    except Exception as e:
        logger.error(f"添加模块失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'添加模块失败: {str(e)}'})


@app.route('/projects/<project_name>/modules/update', methods=['POST'])
def module_update(project_name):
    """更新模块"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    old_name = request.form.get('old_name', '').strip()
    module_name = request.form.get('module_name', '').strip()
    module_desc = request.form.get('module_desc', '').strip()

    logger.info(f"========== 开始更新模块 ==========")
    logger.info(f"项目名称: {project_name}, 旧模块名称: {old_name}, 新模块名称: {module_name}, 描述: {module_desc}")

    if not module_name:
        return jsonify({'success': False, 'error': '模块名称不能为空'})

    if not db_handler:
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 查询模块ID
        mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, old_name))
        if not mod:
            logger.warning(f"模块不存在: {old_name}")
            return jsonify({'success': False, 'error': '模块不存在'})
        mod_id = mod[0]['id']

        # 如果修改了模块名称，检查新名称是否已存在
        if old_name != module_name:
            existing = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
            if existing:
                logger.warning(f"模块已存在: {module_name}")
                return jsonify({'success': False, 'error': '模块已存在'})

        # 更新模块
        db_handler.execute(
            "UPDATE modules SET name = %s, description = %s WHERE id = %s",
            (module_name, module_desc, mod_id)
        )
        logger.info("模块更新成功")

        # 如果修改了模块名称，同步更新相关定时任务
        if old_name != module_name:
            _update_scheduler_jobs_module_name(project_name, old_name, module_name)

        # 清除项目缓存，确保下次获取最新数据
        clear_project_cache()

        return jsonify({'success': True, 'message': '模块更新成功'})
    except Exception as e:
        logger.error(f"更新模块失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'更新模块失败: {str(e)}'})


@app.route('/projects/<project_name>/modules/<module_name>/apis/add', methods=['POST'])
def api_add(project_name, module_name):
    """添加接口"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    module_name = unquote(module_name)
    case_name = request.form.get('case_name', '').strip()
    url = request.form.get('url', '').strip()
    method = request.form.get('method', 'GET').strip()
    headers = request.form.get('headers', '{}').strip()
    data = request.form.get('data', '{}').strip()
    expected = request.form.get('expected', '{}').strip()
    extractions = request.form.get('extractions', '{}').strip()

    logger.info(f"========== 开始添加接口 ==========")
    logger.info(f"项目名称: {project_name}, 模块名称: {module_name}, 接口名称: {case_name}")

    if not case_name:
        return jsonify({'success': False, 'error': '接口名称不能为空'})
    if not url:
        return jsonify({'success': False, 'error': 'URL不能为空'})

    if not db_handler:
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 查询模块ID
        mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
        if not mod:
            logger.warning(f"模块不存在: {module_name}")
            return jsonify({'success': False, 'error': '模块不存在'})
        mod_id = mod[0]['id']

        # 解析JSON数据
        try:
            headers_dict = json.loads(headers) if headers else {}
        except json.JSONDecodeError:
            headers_dict = {}

        try:
            data_dict = json.loads(data) if data else {}
        except json.JSONDecodeError:
            # 纯文本数据，包装为对象保留原始内容
            data_dict = {'_raw': data} if data else {}

        try:
            expected_dict = json.loads(expected) if expected else {}
        except json.JSONDecodeError:
            expected_dict = {}

        try:
            extractions_dict = json.loads(extractions) if extractions else {}
        except json.JSONDecodeError:
            extractions_dict = {}

        # 添加接口
        db_handler.execute(
            "INSERT INTO apis (module_id, case_name, url, method, headers, data, expected, extractions) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
            (mod_id, case_name, url, method, json.dumps(headers_dict, ensure_ascii=False), json.dumps(data_dict, ensure_ascii=False), json.dumps(expected_dict, ensure_ascii=False), json.dumps(extractions_dict, ensure_ascii=False))
        )
        logger.info("接口添加成功")

        return jsonify({'success': True, 'message': '接口添加成功'})
    except Exception as e:
        logger.error(f"添加接口失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'添加接口失败: {str(e)}'})


@app.route('/projects/<project_name>/modules/<module_name>/apis/update/<int:api_id>', methods=['POST'])
def api_update(project_name, module_name, api_id):
    """更新接口"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    module_name = unquote(module_name)
    case_name = request.form.get('case_name', '').strip()
    url = request.form.get('url', '').strip()
    method = request.form.get('method', 'GET').strip()
    headers = request.form.get('headers', '{}').strip()
    data = request.form.get('data', '{}').strip()
    expected = request.form.get('expected', '{}').strip()
    extractions = request.form.get('extractions', '{}').strip()

    logger.info(f"========== 开始更新接口 ==========")
    logger.info(f"项目名称: {project_name}, 模块名称: {module_name}, 接口ID: {api_id}, 接口名称: {case_name}")

    if not case_name:
        return jsonify({'success': False, 'error': '接口名称不能为空'})
    if not url:
        return jsonify({'success': False, 'error': 'URL不能为空'})

    if not db_handler:
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 查询模块ID
        mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
        if not mod:
            logger.warning(f"模块不存在: {module_name}")
            return jsonify({'success': False, 'error': '模块不存在'})
        mod_id = mod[0]['id']

        # 查询接口是否存在
        api = db_handler.query("SELECT * FROM apis WHERE id = %s", (api_id,))
        if not api:
            logger.warning(f"接口不存在: {api_id}")
            return jsonify({'success': False, 'error': '接口不存在'})

        # 解析JSON数据
        try:
            headers_dict = json.loads(headers) if headers else {}
        except json.JSONDecodeError:
            headers_dict = {}

        try:
            data_dict = json.loads(data) if data else {}
        except json.JSONDecodeError:
            # 纯文本数据，包装为对象保留原始内容
            data_dict = {'_raw': data} if data else {}

        try:
            expected_dict = json.loads(expected) if expected else {}
        except json.JSONDecodeError:
            expected_dict = {}

        try:
            extractions_dict = json.loads(extractions) if extractions else {}
        except json.JSONDecodeError:
            extractions_dict = {}

        # 记录旧的接口名称，用于同步更新定时任务
        old_case_name = api[0].get('case_name', '')

        # 更新接口
        db_handler.execute(
            "UPDATE apis SET case_name = %s, url = %s, method = %s, headers = %s, data = %s, expected = %s, extractions = %s WHERE id = %s",
            (case_name, url, method, json.dumps(headers_dict, ensure_ascii=False), json.dumps(data_dict, ensure_ascii=False), json.dumps(expected_dict, ensure_ascii=False), json.dumps(extractions_dict, ensure_ascii=False), api_id)
        )
        logger.info("接口更新成功")

        # 如果接口名称发生变化，同步更新关联的定时任务名称
        if old_case_name != case_name:
            _update_scheduler_jobs_case_name(project_name, module_name, api_id, old_case_name, case_name)

        # 清除项目缓存，确保下次获取最新数据
        clear_project_cache()

        return jsonify({'success': True, 'message': '接口更新成功'})
    except Exception as e:
        logger.error(f"更新接口失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'更新接口失败: {str(e)}'})


@app.route('/projects/<project_name>/modules/<module_name>/apis/delete/<int:api_id>', methods=['POST'])
def apis_delete(project_name, module_name, api_id):
    """删除接口"""
    logger.info(f"========== 开始删除接口 ==========")
    logger.info(f"项目名称: {project_name}, 模块名称: {module_name}, 接口ID: {api_id}")

    if not db_handler:
        logger.error("数据库未连接")
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 查询模块ID
        mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
        if not mod:
            logger.warning(f"模块不存在: {module_name}")
            return jsonify({'success': False, 'error': '模块不存在'})
        mod_id = mod[0]['id']

        # 查询接口是否存在
        api = db_handler.query("SELECT * FROM apis WHERE id = %s", (api_id,))
        if not api:
            logger.warning(f"接口不存在: {api_id}")
            return jsonify({'success': False, 'error': '接口不存在'})

        # 删除接口
        db_handler.execute("DELETE FROM apis WHERE id = %s", (api_id,))
        logger.info("接口删除成功")

        return jsonify({'success': True, 'message': '接口删除成功'})
    except Exception as e:
        logger.error(f"删除接口失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'接口删除失败: {str(e)}'})


# 路由：获取单个接口数据
@app.route('/projects/<project_name>/modules/<module_name>/apis')
def apis_list(project_name, module_name):
    """获取模块下的接口列表"""
    logger.info(f"========== 开始获取接口列表 ==========")
    logger.info(f"项目名称: {project_name}, 模块名称: {module_name}")

    if not db_handler:
        logger.error("数据库未连接")
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 查询模块ID
        mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
        if not mod:
            logger.warning(f"模块不存在: {module_name}")
            return jsonify({'success': False, 'error': '模块不存在'})
        mod_id = mod[0]['id']

        # 查询接口列表
        apis = db_handler.query("SELECT * FROM apis WHERE module_id = %s ORDER BY id", (mod_id,))
        logger.info(f"获取到 {len(apis)} 个接口")

        # 返回接口列表
        api_list = []
        for api in apis:
            api_list.append({
                'id': api.get('id'),
                'case_name': api.get('case_name'),
                'url': api.get('url'),
                'method': api.get('method')
            })

        return jsonify({'success': True, 'apis': api_list})
    except Exception as e:
        logger.error(f"获取接口列表失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'获取接口列表失败: {str(e)}'})


@app.route('/projects/<project_name>/modules/<module_name>/apis/get/<int:api_id>')
def apis_get(project_name, module_name, api_id):
    """获取单个接口数据"""
    logger.info(f"========== 开始获取接口数据 ==========")
    logger.info(f"项目名称: {project_name}, 模块名称: {module_name}, 接口ID: {api_id}")

    if not db_handler:
        logger.error("数据库未连接")
        return jsonify({'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 查询模块ID
        mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
        if not mod:
            logger.warning(f"模块不存在: {module_name}")
            return jsonify({'error': '模块不存在'})
        mod_id = mod[0]['id']

        # 查询接口
        api = db_handler.query("SELECT * FROM apis WHERE id = %s", (api_id,))
        if not api:
            logger.warning(f"接口不存在: {api_id}")
            return jsonify({'error': '接口不存在'})

        api_data = api[0]
        logger.info(f"获取接口数据: {api_data}")

        # 返回接口数据
        return jsonify({
            'id': api_data.get('id'),
            'api_index': api_data.get('id'),
            'case_name': api_data.get('case_name'),
            'url': api_data.get('url'),
            'method': api_data.get('method'),
            'headers': json.loads(api_data.get('headers', '{}')),
            'data': _unwrap_raw(json.loads(api_data.get('data', '{}'))),
            'expected': json.loads(api_data.get('expected', '{}')),
            'extractions': json.loads(api_data.get('extractions', '{}')),
            'project_name': project_name,
            'module_name': module_name
        })
    except Exception as e:
        logger.error(f"获取接口数据失败: {e}", exc_info=True)
        return jsonify({'error': f'获取接口数据失败: {str(e)}'})


# 路由：添加模块
@app.route('/projects/<project_name>/modules/add', methods=['POST'])
def add_module(project_name):
    """添加模块"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    module_name = request.form.get('module_name', '').strip()
    module_desc = request.form.get('module_desc', '').strip()

    if not module_name:
        return jsonify({'success': False, 'error': '模块名称不能为空'})

    if not db_handler:
        return jsonify({'success': False, 'error': '数据库未连接'})

    try:
        db_handler._ensure_connection()

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            return jsonify({'success': False, 'error': '项目不存在'})
        proj_id = proj[0]['id']

        # 检查模块是否已存在
        existing = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s",
                                    (proj_id, module_name))
        if existing:
            return jsonify({'success': False, 'error': '模块已存在'})

        # 添加模块
        db_handler.execute(
            "INSERT INTO modules (project_id, name, description) VALUES (%s, %s, %s)",
            (proj_id, module_name, module_desc)
        )

        return jsonify({'success': True, 'message': '模块添加成功'})
    except Exception as e:
        logger.error(f"添加模块失败: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'添加模块失败: {str(e)}'})


# 路由：添加API（旧版，基于YAML文件）
@app.route('/api/add', methods=['POST'])
def api_add_legacy():
    """添加API"""
    api_name = request.form.get('api_name')
    case_name = request.form.get('case_name')
    url = request.form.get('url')
    method = request.form.get('method')
    headers = request.form.get('headers', '{}')
    data = request.form.get('data', '{}')
    expected = request.form.get('expected', '{}')

    try:
        # 解析JSON数据
        headers_dict = json.loads(headers) if headers else {}
        try:
            data_dict = json.loads(data) if data else {}
        except Exception:
            # 如果不是有效JSON，包装为对象保留原始内容
            data_dict = {'_raw': data} if data else {}
        expected_dict = json.loads(expected) if expected else {}

        # 获取测试数据
        test_data = get_test_data()

        # 如果API名称不存在，则创建
        if api_name not in test_data:
            test_data[api_name] = []

        # 添加测试用例
        test_data[api_name].append({
            "case_name": case_name,
            "url": url,
            "method": method,
            "headers": headers_dict,
            "data": data_dict,
            "expected": expected_dict
        })

        # 保存测试数据
        if save_test_data(test_data):
            flash('API添加成功', 'success')
        else:
            flash('API添加失败', 'error')
    except Exception as e:
        logger.error(f"添加API失败: {e}")
        flash(f'API添加失败: {str(e)}', 'error')

    return redirect(url_for('index'))


# 路由：删除API
@app.route('/api/delete/<api_name>/<int:case_index>')
def api_delete(api_name, case_index):
    """删除API"""
    test_data = get_test_data()

    if api_name in test_data and 0 <= case_index < len(test_data[api_name]):
        test_data[api_name].pop(case_index)

        # 如果API名称下没有测试用例了，则删除该API名称
        if not test_data[api_name]:
            del test_data[api_name]

        # 保存测试数据
        if save_test_data(test_data):
            flash('API删除成功', 'success')
        else:
            flash('API删除失败', 'error')
    else:
        flash('API不存在或索引错误', 'error')

    return redirect(url_for('index'))


# 路由：编辑API
@app.route('/api/edit/<api_name>/<int:case_index>', methods=['GET', 'POST'])
def api_edit(api_name, case_index):
    """编辑API"""
    test_data = get_test_data()

    if request.method == 'GET':
        # 获取API数据
        if api_name in test_data and 0 <= case_index < len(test_data[api_name]):
            api_data = test_data[api_name][case_index]
            return jsonify({
                "api_name": api_name,
                "case_index": case_index,
                "case_name": api_data.get("case_name", ""),
                "url": api_data.get("url", ""),
                "method": api_data.get("method", "GET"),
                "headers": api_data.get("headers", {}),
                "data": api_data.get("data", {}),
                "expected": api_data.get("expected", {})
            })
        else:
            return jsonify({"error": "API不存在或索引错误"}), 404
    else:
        # 更新API数据
        case_name = request.form.get('case_name')
        url = request.form.get('url')
        method = request.form.get('method')
        headers = request.form.get('headers', '{}')
        data = request.form.get('data', '{}')
        expected = request.form.get('expected', '{}')

        try:
            # 解析JSON数据
            headers_dict = json.loads(headers) if headers else {}
            data_dict = json.loads(data) if data else {}
            expected_dict = json.loads(expected) if expected else {}

            # 更新测试用例
            if api_name in test_data and 0 <= case_index < len(test_data[api_name]):
                test_data[api_name][case_index] = {
                    "case_name": case_name,
                    "url": url,
                    "method": method,
                    "headers": headers_dict,
                    "data": data_dict,
                    "expected": expected_dict
                }

                # 保存测试数据
                if save_test_data(test_data):
                    flash('API更新成功', 'success')
                else:
                    flash('API更新失败', 'error')
            else:
                flash('API不存在或索引错误', 'error')
        except Exception as e:
            logger.error(f"更新API失败: {e}")
            flash(f'API更新失败: {str(e)}', 'error')

        return redirect(url_for('index'))


# 路由：调试API
@app.route('/api/debug', methods=['POST'])
def api_debug():
    """调试API"""
    api_data = request.json
    api_data['source'] = '调试'
    result = execute_api_test(api_data)
    return jsonify(result)


# 路由：定时任务列表
@app.route('/scheduler/list')
def scheduler_list():
    """定时任务列表（无缓存，确保实时性）"""
    # 从数据库读取排序信息
    db_sort_orders = {}
    if db_handler:
        try:
            db_handler._ensure_connection()
            sort_rows = db_handler.query("SELECT id, sort_order FROM scheduler_jobs")
            for row in sort_rows:
                db_sort_orders[row['id']] = row.get('sort_order', 0) or 0
        except Exception:
            pass

    jobs = []
    for job in scheduler.get_jobs():
        # 从job.args中提取参数
        project_name = job.args[0] if len(job.args) > 0 else ''
        module_name = job.args[1] if len(job.args) > 1 else ''
        api_id = job.args[2] if len(job.args) > 2 else 0

        # 从trigger中提取cron表达式
        cron_expression = _extract_cron_expression(job.trigger)

        sort_order = db_sort_orders.get(job.id, 0)

        # 处理next_run_time，确保正确处理时区
        next_run_time_str = None
        if job.next_run_time:
            try:
                # 如果next_run_time有时区信息，转换为本地时间
                if job.next_run_time.tzinfo is not None:
                    next_run_time_str = job.next_run_time.astimezone().strftime('%Y-%m-%d %H:%M:%S')
                else:
                    next_run_time_str = job.next_run_time.strftime('%Y-%m-%d %H:%M:%S')
            except Exception as e:
                logger.error(f"格式化next_run_time失败: {e}")
                next_run_time_str = None

        jobs.append({
            'id': job.id,
            'name': job.name,
            'next_run_time': next_run_time_str,
            'trigger': str(job.trigger),
            'project_name': project_name,
            'module_name': module_name,
            'api_id': api_id,
            'cron_expression': cron_expression,
            'sort_order': sort_order
        })

    # 按 sort_order 排序，sort_order 相同的按 id 排序
    jobs.sort(key=lambda x: (x['sort_order'], x['id']))

    return jsonify(jobs)


# 路由：添加定时任务
@app.route('/scheduler/add', methods=['POST'])
def scheduler_add():
    """添加定时任务"""
    project_name = request.form.get('project_name')
    module_name = request.form.get('module_name')
    api_id_str = request.form.get('case_index') or request.form.get('api_id', '0')
    api_id = int(api_id_str) if api_id_str and api_id_str != 'undefined' else 0
    cron_expression = request.form.get('cron_expression')

    try:
        # 对URL编码的参数进行解码
        from urllib.parse import unquote
        if project_name:
            project_name = unquote(project_name)
        if module_name:
            module_name = unquote(module_name)

        # 从数据库中获取接口数据
        api_data = None
        if project_name and module_name and api_id and db_handler:
            try:
                db_handler._ensure_connection()
                # 查询项目ID
                proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
                if proj:
                    proj_id = proj[0]['id']
                    # 查询模块ID
                    mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
                    if mod:
                        mod_id = mod[0]['id']
                        # 查询接口
                        api = db_handler.query("SELECT * FROM apis WHERE id = %s", (api_id,))
                        if api:
                            api_data = api[0]
            except Exception as e:
                logger.error(f"获取接口数据失败: {e}", exc_info=True)

        if not api_data:
            return jsonify({'success': False, 'message': 'API不存在或ID错误'}), 400

        # 校验cron表达式格式
        cron_fields = cron_expression.strip().split()
        if len(cron_fields) != 5:
            return jsonify({'success': False, 'message': f'Cron表达式格式错误：需要5个字段（分 时 日 月 周），当前只有{len(cron_fields)}个字段。示例：0 * * * * 表示每小时执行'}), 400

        # 创建任务ID
        job_id = f"{project_name}_{module_name}_{api_id}_{int(time.time())}"

        # 计算新任务的排序值（当前模块最大sort_order + 1）
        new_sort_order = 0
        if db_handler:
            try:
                max_sort = db_handler.query(
                    "SELECT MAX(sort_order) as max_sort FROM scheduler_jobs WHERE project_name = %s AND module_name = %s",
                    (project_name, module_name)
                )
                if max_sort and max_sort[0].get('max_sort') is not None:
                    new_sort_order = max_sort[0]['max_sort'] + 1
            except Exception:
                pass

        # 添加定时任务
        scheduler.add_job(
            func=scheduled_test_job,
            trigger=CronTrigger.from_crontab(cron_expression),
            id=job_id,
            args=[project_name, module_name, api_id],
            name=f"{project_name}/{module_name} - {api_data.get('case_name', '未命名')}"
        )

        # 触发一次next_run_time的计算
        job = scheduler.get_job(job_id)
        if job and job.next_run_time:
            logger.info(f"定时任务 {job_id} 下次执行时间: {job.next_run_time.strftime('%Y-%m-%d %H:%M:%S')}")

        # 直接插入数据库中的定时任务
        if db_handler:
            try:
                db_handler._ensure_connection()
                db_handler.execute(
                    """INSERT INTO scheduler_jobs (id, name, project_name, module_name, case_index, cron_expression, sort_order)
                       VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                    (
                        job_id,
                        f"{project_name}/{module_name} - {api_data.get('case_name', '未命名')}",
                        project_name,
                        module_name,
                        api_id,
                        cron_expression,
                        new_sort_order
                    )
                )
                logger.info(f"已添加定时任务到数据库: {job_id}, CRON表达式: {cron_expression}")
            except Exception as e:
                logger.error(f"添加定时任务到数据库失败: {e}")

        return jsonify({'success': True, 'message': '定时任务添加成功'})
    except ValueError as e:
        logger.error(f"添加定时任务失败(Cron表达式无效): {e}")
        return jsonify({'success': False, 'message': f'Cron表达式无效: {str(e)}'}), 400
    except Exception as e:
        logger.error(f"添加定时任务失败: {e}")
        return jsonify({'success': False, 'message': f'添加定时任务失败: {str(e)}'}), 500


# 路由：更新定时任务
@app.route('/scheduler/update', methods=['POST'])
def scheduler_update():
    """更新定时任务"""
    job_id = request.form.get('job_id')
    project_name = request.form.get('project_name')
    module_name = request.form.get('module_name')
    api_id_str = request.form.get('case_index') or request.form.get('api_id', '0')
    api_id = int(api_id_str) if api_id_str and api_id_str != 'undefined' else 0
    cron_expression = request.form.get('cron_expression')

    try:
        # 校验cron表达式格式
        cron_fields = cron_expression.strip().split()
        if len(cron_fields) != 5:
            return jsonify({'success': False, 'message': f'Cron表达式格式错误：需要5个字段（分 时 日 月 周），当前只有{len(cron_fields)}个字段。示例：0 * * * * 表示每小时执行'}), 400

        # 检查原任务是否存在
        existing_job = scheduler.get_job(job_id)
        if not existing_job:
            return jsonify({'success': False, 'message': '定时任务不存在，可能已被删除'}), 404

        # 获取原任务的名称（保持不变）
        job_name = existing_job.name

        # 删除旧任务
        scheduler.remove_job(job_id)

        # 创建新任务（使用原job_id）
        scheduler.add_job(
            func=scheduled_test_job,
            trigger=CronTrigger.from_crontab(cron_expression),
            id=job_id,
            args=[project_name, module_name, api_id],
            name=job_name
        )

        # 触发一次next_run_time的计算
        job = scheduler.get_job(job_id)
        if job and job.next_run_time:
            logger.info(f"定时任务 {job_id} 下次执行时间: {job.next_run_time.strftime('%Y-%m-%d %H:%M:%S')}")

        # 直接更新数据库中的CRON表达式，而不是清空整个表
        if db_handler:
            try:
                db_handler._ensure_connection()
                db_handler.execute(
                    "UPDATE scheduler_jobs SET cron_expression = %s WHERE id = %s",
                    (cron_expression, job_id)
                )
                logger.info(f"已更新数据库中的定时任务CRON表达式: {job_id} -> {cron_expression}")
            except Exception as e:
                logger.error(f"更新数据库中的定时任务CRON表达式失败: {e}")

        return jsonify({'success': True, 'message': '定时任务更新成功'})
    except ValueError as e:
        logger.error(f"更新定时任务失败(Cron表达式无效): {e}")
        return jsonify({'success': False, 'message': f'Cron表达式无效: {str(e)}'}), 400
    except Exception as e:
        logger.error(f"更新定时任务失败: {e}")
        return jsonify({'success': False, 'message': f'更新定时任务失败: {str(e)}'}), 500


# 路由：删除定时任务
@app.route('/scheduler/delete/<job_id>')
def scheduler_delete(job_id):
    """删除定时任务"""
    try:
        scheduler.remove_job(job_id)

        # 直接从数据库中删除定时任务
        if db_handler:
            try:
                db_handler._ensure_connection()
                db_handler.execute("DELETE FROM scheduler_jobs WHERE id = %s", (job_id,))
                logger.info(f"已从数据库中删除定时任务: {job_id}")
            except Exception as e:
                logger.error(f"从数据库中删除定时任务失败: {e}")

        return jsonify({'success': True, 'message': '定时任务删除成功'})
    except Exception as e:
        logger.error(f"删除定时任务失败: {e}")
        return jsonify({'success': False, 'error': f'删除定时任务失败: {str(e)}'})


# 路由：定时任务排序移动
@app.route('/scheduler/move', methods=['POST'])
def scheduler_move():
    """移动定时任务排序（上移/下移）"""
    try:
        job_id = request.form.get('job_id')
        direction = request.form.get('direction')  # 'up' 或 'down'

        if not job_id or direction not in ('up', 'down'):
            return jsonify({'success': False, 'message': '参数错误'}), 400

        if not db_handler:
            return jsonify({'success': False, 'message': '数据库未连接'}), 500

        db_handler._ensure_connection()

        # 获取当前job的排序和所属模块
        current_job = db_handler.query("SELECT sort_order, project_name, module_name FROM scheduler_jobs WHERE id = %s", (job_id,))
        if not current_job:
            return jsonify({'success': False, 'message': '任务不存在'}), 404

        current_sort = current_job[0].get('sort_order', 0) or 0
        project_name = current_job[0]['project_name']
        module_name = current_job[0]['module_name']

        # 获取同模块下所有任务，按sort_order排序
        module_jobs = db_handler.query(
            "SELECT id, sort_order FROM scheduler_jobs WHERE project_name = %s AND module_name = %s ORDER BY sort_order, id",
            (project_name, module_name)
        )

        if len(module_jobs) <= 1:
            return jsonify({'success': True, 'message': '只有一个任务，无需移动'})

        # 重新分配连续的sort_order（避免排序值不连续）
        for i, mj in enumerate(module_jobs):
            if mj['sort_order'] != i:
                db_handler.execute("UPDATE scheduler_jobs SET sort_order = %s WHERE id = %s", (i, mj['id']))

        # 重新获取排序后的列表
        module_jobs = db_handler.query(
            "SELECT id, sort_order FROM scheduler_jobs WHERE project_name = %s AND module_name = %s ORDER BY sort_order, id",
            (project_name, module_name)
        )

        # 找到当前job在列表中的索引
        current_index = -1
        for i, mj in enumerate(module_jobs):
            if mj['id'] == job_id:
                current_index = i
                break

        if direction == 'up' and current_index <= 0:
            return jsonify({'success': True, 'message': '已经在最前面'})
        if direction == 'down' and current_index >= len(module_jobs) - 1:
            return jsonify({'success': True, 'message': '已经在最后面'})

        # 交换当前job和目标job的sort_order
        swap_index = current_index - 1 if direction == 'up' else current_index + 1
        swap_job_id = module_jobs[swap_index]['id']

        db_handler.execute("UPDATE scheduler_jobs SET sort_order = %s WHERE id = %s", (swap_index, job_id))
        db_handler.execute("UPDATE scheduler_jobs SET sort_order = %s WHERE id = %s", (current_index, swap_job_id))

        # 清除缓存
        scheduler_cache.clear()

        return jsonify({'success': True, 'message': '移动成功'})
    except Exception as e:
        logger.error(f"移动定时任务排序失败: {e}")
        return jsonify({'success': False, 'message': f'移动失败: {str(e)}'}), 500


# 路由：获取统计页面的项目列表
@app.route('/statistics/projects')
@cache_result('statistics_projects', stats_cache, timeout=30)  # 缓存30秒
def statistics_projects():
    """获取统计数据中所有项目列表（SQL直接查询）"""

    try:
        if not db_handler:
            return jsonify({'success': True, 'projects': []})
        db_handler._ensure_connection()
        rows = db_handler.query("SELECT DISTINCT project FROM test_statistics WHERE project IS NOT NULL AND project != '' ORDER BY project")
        projects = [r['project'] for r in rows]
        return jsonify({'success': True, 'projects': projects})
    except Exception as e:
        logger.error(f"获取项目列表失败: {e}")
        return jsonify({'success': True, 'projects': []})


# 路由：获取统计页面的模块列表
@app.route('/statistics/modules')
@cache_result('statistics_modules', stats_cache, timeout=30)  # 缓存30秒
def statistics_modules():
    """获取指定项目的模块列表（SQL直接查询）"""

    project = request.args.get('project', '')
    try:
        if not db_handler:
            return jsonify({'success': True, 'modules': []})
        db_handler._ensure_connection()
        if project:
            rows = db_handler.query("SELECT DISTINCT module FROM test_statistics WHERE project = %s AND module IS NOT NULL AND module != '' ORDER BY module", (project,))
        else:
            rows = db_handler.query("SELECT DISTINCT module FROM test_statistics WHERE module IS NOT NULL AND module != '' ORDER BY module")
        modules = [r['module'] for r in rows]
        return jsonify({'success': True, 'modules': modules})
    except Exception as e:
        logger.error(f"获取模块列表失败: {e}")
        return jsonify({'success': True, 'modules': []})


# 路由：获取统计数据列表
@app.route('/statistics/list')
def statistics_list():
    """获取统计数据列表，支持筛选和分页（SQL层筛选+分页，高性能）"""
    if not db_handler:
        return jsonify({'records': [], 'total': 0, 'page': 1, 'page_size': 20,
                        'total_pages': 0, 'data_source': 'DATABASE', 'summary': {}})

    # 获取筛选参数
    project = request.args.get('project', '')
    module = request.args.get('module', '')
    method = request.args.get('method', '')
    status = request.args.get('status', '')
    assertion = request.args.get('assertion', '')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    keyword = request.args.get('keyword', '')
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))

    # 构建SQL WHERE条件
    conditions = []
    params = []

    if project:
        conditions.append("project = %s")
        params.append(project)
    if module:
        conditions.append("module = %s")
        params.append(module)
    if method:
        conditions.append("method = %s")
        params.append(method)
    if status:
        if status == '2xx':
            conditions.append("status_code >= 200 AND status_code < 300")
        elif status == '4xx':
            conditions.append("status_code >= 400 AND status_code < 500")
        elif status == '5xx':
            conditions.append("status_code >= 500 AND status_code < 600")
    if assertion:
        conditions.append("assertion_passed = %s")
        params.append(assertion == 'passed')
    if date_start:
        conditions.append("timestamp >= %s")
        params.append(date_start)
    if date_end:
        conditions.append("timestamp <= %s")
        params.append(date_end + ' 23:59:59')
    if keyword:
        conditions.append("(LOWER(url) LIKE %s OR LOWER(project) LIKE %s OR LOWER(module) LIKE %s)")
        kw_like = f'%{keyword.lower()}%'
        params.extend([kw_like, kw_like, kw_like])

    where_clause = f" WHERE {' AND '.join(conditions)}" if conditions else ""

    try:
        db_handler._ensure_connection()

        # 1. 统计概览（使用SQL聚合，无需加载全量数据）
        summary_sql = f"""SELECT
            COUNT(*) as total_count,
            SUM(CASE WHEN status_code >= 200 AND status_code < 300 THEN 1 ELSE 0 END) as success_count,
            ROUND(AVG(response_time), 2) as avg_time,
            ROUND(MIN(response_time), 2) as min_time,
            ROUND(MAX(response_time), 2) as max_time,
            SUM(assertion_count) as assertion_total,
            SUM(assertion_passed_count) as assertion_passed_total
        FROM test_statistics{where_clause}"""
        summary_row = db_handler.query(summary_sql, tuple(params) if params else None)
        summary_data = summary_row[0] if summary_row else {}

        total = summary_data.get('total_count', 0) or 0
        success_count = summary_data.get('success_count', 0) or 0
        fail_count = total - success_count
        avg_time = summary_data.get('avg_time', 0) or 0
        min_time = summary_data.get('min_time', 0) or 0
        max_time = summary_data.get('max_time', 0) or 0
        assertion_total = summary_data.get('assertion_total', 0) or 0
        assertion_passed_total = summary_data.get('assertion_passed_total', 0) or 0

        # 2. 状态码分布（SQL聚合，兼容 ONLY_FULL_GROUP_BY）
        status_dist_sql = f"""SELECT
            CONCAT(FLOOR(status_code / 100), 'xx') as status_group,
            COUNT(*) as cnt
        FROM test_statistics{where_clause}
        GROUP BY CONCAT(FLOOR(status_code / 100), 'xx')"""
        status_dist_rows = db_handler.query(status_dist_sql, tuple(params) if params else None)
        status_dist = {r['status_group']: r['cnt'] for r in status_dist_rows} if status_dist_rows else {}

        # 3. 请求方法分布（SQL聚合）
        method_dist_sql = f"""SELECT method, COUNT(*) as cnt
        FROM test_statistics{where_clause}
        GROUP BY method"""
        method_dist_rows = db_handler.query(method_dist_sql, tuple(params) if params else None)
        method_dist = {r['method']: r['cnt'] for r in method_dist_rows} if method_dist_rows else {}

        # 4. 响应时间趋势（最近50条，SQL LIMIT）
        trend_sql = f"""SELECT timestamp as time, response_time as value
        FROM test_statistics{where_clause}
        ORDER BY id DESC LIMIT 50"""
        trend_rows = db_handler.query(trend_sql, tuple(params) if params else None)
        trend_data = list(reversed(trend_rows)) if trend_rows else []

        # 5. 分页数据（SQL LIMIT OFFSET）
        total_pages = max(1, (total + page_size - 1) // page_size)
        page = min(page, total_pages)
        offset = (page - 1) * page_size

        page_sql = f"""SELECT id, method, url, status_code, response_time,
            assertion_passed, assertion_count, assertion_passed_count,
            source, project, module, case_name, request_headers,
            request_body, response_headers, response_body,
            timestamp, error, created_at
        FROM test_statistics{where_clause}
        ORDER BY id DESC
        LIMIT %s OFFSET %s"""
        page_params = list(params) + [page_size, offset]
        page_data = db_handler.query(page_sql, tuple(page_params))

        # 批量获取当前页的断言结果
        if page_data:
            stat_ids = [int(r['id']) for r in page_data if r.get('id')]
            if stat_ids:
                placeholders = ','.join(['%s'] * len(stat_ids))
                all_assertions = db_handler.query(
                    f"SELECT * FROM assertion_results WHERE statistic_id IN ({placeholders})",
                    tuple(stat_ids)
                )
                assertion_map = {}
                for ar in all_assertions:
                    sid = ar.get('statistic_id')
                    if sid not in assertion_map:
                        assertion_map[sid] = []
                    assertion_map[sid].append(ar)
            else:
                assertion_map = {}

            for r in page_data:
                # 映射 timestamp
                if 'timestamp' not in r or not r['timestamp']:
                    r['timestamp'] = str(r.get('created_at', '')) if r.get('created_at') else ''
                # 确保 error 字段存在
                if 'error' not in r:
                    r['error'] = ''
                # 断言结果
                statistic_id = int(r['id']) if r['id'] else 0
                assertions = assertion_map.get(statistic_id, [])
                # 确保passed字段是布尔类型
                for assertion in assertions:
                    if 'passed' in assertion:
                        assertion['passed'] = bool(assertion['passed'])
                r['assertion_results'] = assertions

        return jsonify({
            'records': page_data,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'data_source': 'DATABASE',
            'summary': {
                'total_count': total,
                'success_count': success_count,
                'fail_count': fail_count,
                'avg_time': float(avg_time),
                'min_time': float(min_time),
                'max_time': float(max_time),
                'success_rate': round(success_count / total * 100, 1) if total else 0,
                'assertion_rate': round(assertion_passed_total / assertion_total * 100, 1) if assertion_total else 0,
                'status_dist': status_dist,
                'method_dist': method_dist,
                'trend_data': trend_data
            }
        })
    except Exception as e:
        logger.error(f"获取统计数据列表失败: {e}")
        return jsonify({'records': [], 'total': 0, 'page': 1, 'page_size': page_size,
                        'total_pages': 0, 'data_source': 'DATABASE', 'summary': {},
                        'error': str(e)})


# 路由：导出统计数据
@app.route('/statistics/export')
def statistics_export():
    """导出当前筛选的统计数据为Excel文件"""
    data = get_statistics_data()

    # 获取筛选参数（与statistics_list相同的筛选逻辑）
    project = request.args.get('project', '')
    module = request.args.get('module', '')
    method = request.args.get('method', '')
    status = request.args.get('status', '')
    assertion = request.args.get('assertion', '')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    keyword = request.args.get('keyword', '')

    # 筛选
    filtered = data
    if project:
        filtered = [r for r in filtered if r.get('project') == project]
    if module:
        filtered = [r for r in filtered if r.get('module') == module]
    if method:
        filtered = [r for r in filtered if r.get('method') == method]
    if status:
        filtered = [r for r in filtered if r.get('status_code') and (
            (status == '2xx' and 200 <= r['status_code'] < 300) or
            (status == '4xx' and 400 <= r['status_code'] < 500) or
            (status == '5xx' and 500 <= r['status_code'] < 600)
        )]
    if assertion:
        filtered = [r for r in filtered if r.get('assertion_passed') == (assertion == 'passed')]
    if date_start:
        filtered = [r for r in filtered if r.get('timestamp', '') >= date_start]
    if date_end:
        filtered = [r for r in filtered if r.get('timestamp', '') <= date_end + ' 23:59:59']
    if keyword:
        kw = keyword.lower()
        filtered = [r for r in filtered if kw in r.get('url', '').lower() or kw in r.get('project', '').lower() or kw in r.get('module', '').lower()]

    # 倒序（最新的在前）
    filtered.reverse()

    # 使用openpyxl生成Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '统计数据'

    # 表头定义（与页面表格列顺序一致：ID、项目、模块、接口名称、来源、方法、URL、状态码、响应时间、断言、执行时间）
    headers = ['ID', '项目', '模块', '接口名称', '来源', '请求方法', 'URL', '状态码', '响应时间(ms)', '断言结果', '执行时间']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    cell_font = Font(name='微软雅黑', size=9)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 状态码颜色填充
    success_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    danger_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

    for row_idx, r in enumerate(filtered, 2):
        values = [
            r.get('id', ''),           # 1. ID
            r.get('project', ''),       # 2. 项目
            r.get('module', ''),        # 3. 模块
            r.get('case_name', ''),     # 4. 接口名称
            r.get('source', ''),        # 5. 来源
            r.get('method', ''),        # 6. 请求方法
            r.get('url', ''),           # 7. URL
            r.get('status_code', ''),   # 8. 状态码
            r.get('response_time', ''), # 9. 响应时间
            '通过' if r.get('assertion_passed') else '失败',  # 10. 断言结果
            r.get('timestamp', '')      # 11. 执行时间
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.border = thin_border

            # 状态码列着色（第8列）
            if col == 8 and isinstance(value, int):
                if 200 <= value < 300:
                    cell.fill = success_fill
                    cell.font = Font(name='微软雅黑', size=9, color='2E7D32')
                elif 400 <= value < 500:
                    cell.fill = warning_fill
                    cell.font = Font(name='微软雅黑', size=9, color='F57F17')
                elif value >= 500:
                    cell.fill = danger_fill
                    cell.font = Font(name='微软雅黑', size=9, color='C62828')
                cell.alignment = center_alignment
            elif col == 10:  # 断言结果列（第10列）
                if value == '通过':
                    cell.fill = success_fill
                    cell.font = Font(name='微软雅黑', size=9, color='2E7D32')
                else:
                    cell.fill = danger_fill
                    cell.font = Font(name='微软雅黑', size=9, color='C62828')
                cell.alignment = center_alignment
            elif col in (1, 5, 6, 9):  # ID、来源、方法、响应时间居中
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment

    # 设置列宽
    column_widths = {'A': 6, 'B': 15, 'C': 15, 'D': 18, 'E': 10, 'F': 10, 'G': 50, 'H': 10, 'I': 14, 'J': 10, 'K': 20}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 设置自动筛选
    ws.auto_filter.ref = f'A1:K{len(filtered) + 1}'

    # 保存到内存
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    filename = f'统计数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# 路由：获取单条统计详情
@app.route('/statistics/detail')
@app.route('/statistics/detail/<int:record_id>')
def statistics_detail(record_id=None):
    """获取单条统计详情"""
    # 支持路径参数和查询参数两种方式
    if record_id is None:
        try:
            record_id = int(request.args.get('id', 0))
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': '无效的ID参数'}), 400

    try:
        if not db_handler:
            return jsonify({'success': False, 'error': '数据库不可用'}), 500
        db_handler._ensure_connection()
        stat = db_handler.query(
            """SELECT id, method, url, status_code, response_time,
               assertion_passed, assertion_count, assertion_passed_count,
               source, project, module, case_name, request_headers,
               request_body, response_headers, response_body,
               timestamp, error, created_at
               FROM test_statistics WHERE id = %s""",
            (record_id,)
        )
        if stat and len(stat) > 0:
            r = stat[0]
            # 映射 timestamp
            if 'timestamp' not in r or not r['timestamp']:
                r['timestamp'] = str(r.get('created_at', '')) if r.get('created_at') else ''
            if 'error' not in r:
                r['error'] = ''
            # 获取断言结果
            assertion_results = db_handler.query(
                "SELECT * FROM assertion_results WHERE statistic_id = %s",
                (record_id,)
            )
            # 确保passed字段是布尔类型
            for assertion in assertion_results:
                if 'passed' in assertion:
                    assertion['passed'] = bool(assertion['passed'])
            r['assertion_results'] = assertion_results
            return jsonify({'success': True, 'detail': r})
        return jsonify({'success': False, 'error': '记录不存在'}), 404
    except Exception as e:
        logger.error(f"查询统计详情失败: {e}")
        return jsonify({'success': False, 'error': f'查询失败: {str(e)}'}), 500


# 路由：获取项目的环境配置
@app.route('/projects/<project_name>/env/list')
def project_env_list(project_name):
    """获取项目的环境列表和当前环境"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    projects_data = get_projects()
    if project_name not in projects_data:
        return jsonify({'success': False, 'error': '项目不存在'}), 404

    project = projects_data[project_name]
    envs = project.get('envs', {})
    current_env = project.get('current_env', '')

    env_list = []
    for env_name, env_data in envs.items():
        if isinstance(env_data, dict):
            env_list.append({
                'name': env_name,
                'base_url': env_data.get('base_url', '')
            })

    # 获取当前环境的base_url
    current_base_url = ''
    if current_env and current_env in envs and isinstance(envs[current_env], dict):
        current_base_url = envs[current_env].get('base_url', '')

    return jsonify({
        'success': True,
        'current_env': current_env,
        'env_list': env_list,
        'base_url': current_base_url
    })


# 路由：切换项目环境
@app.route('/projects/<project_name>/env/switch', methods=['POST'])
def project_env_switch(project_name):
    """切换项目的当前环境"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    env_name = request.form.get('env_name')
    if not env_name:
        return jsonify({'success': False, 'error': '环境名称不能为空'})

    try:
        db_handler._ensure_connection()
        # 获取项目ID和当前环境
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        project_id = proj[0]['id']

        # 检查环境是否存在
        env = db_handler.query("SELECT name, base_url FROM environments WHERE project_id = %s AND name = %s", (project_id, env_name))
        if not env:
            return jsonify({'success': False, 'error': f'环境不存在: {env_name}'})

        # 更新当前环境
        db_handler.execute("UPDATE projects SET current_env = %s WHERE id = %s", (env_name, project_id))

        base_url = env[0].get('base_url', '')
        return jsonify({
            'success': True,
            'message': f'环境已切换为: {env_name}',
            'current_env': env_name,
            'base_url': base_url
        })
    except Exception as e:
        logger.error(f"切换环境失败: {e}")
        return jsonify({'success': False, 'error': f'切换失败: {str(e)}'})


# 路由：添加/编辑项目环境
@app.route('/projects/<project_name>/env/save', methods=['POST'])
def project_env_save(project_name):
    """添加或编辑项目的环境配置"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    env_name = request.form.get('env_name', '').strip()
    base_url = request.form.get('base_url', '').strip()

    if not env_name:
        return jsonify({'success': False, 'error': '环境名称不能为空'})
    if not base_url:
        return jsonify({'success': False, 'error': '环境域名不能为空'})

    try:
        db_handler._ensure_connection()
        # 获取项目ID
        proj = db_handler.query("SELECT id, current_env FROM projects WHERE name = %s", (project_name,))
        if not proj:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        project_id = proj[0]['id']
        current_env = proj[0].get('current_env', '') or ''

        # 插入或更新环境（利用UNIQUE KEY uk_project_env）
        db_handler.execute(
            "INSERT INTO environments (project_id, name, base_url) VALUES (%s, %s, %s) "
            "ON DUPLICATE KEY UPDATE base_url = %s",
            (project_id, env_name, base_url, base_url)
        )

        # 如果是第一个环境，自动设为当前环境
        if not current_env:
            db_handler.execute("UPDATE projects SET current_env = %s WHERE id = %s", (env_name, project_id))

        return jsonify({'success': True, 'message': f'环境 {env_name} 保存成功'})
    except Exception as e:
        logger.error(f"保存环境失败: {e}")
        return jsonify({'success': False, 'error': f'保存失败: {str(e)}'})


# 路由：删除项目环境
@app.route('/projects/<project_name>/env/delete/<env_name>', methods=['POST'])
def project_env_delete(project_name, env_name):
    """删除项目的环境配置"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    env_name = unquote(env_name)

    try:
        db_handler._ensure_connection()
        proj = db_handler.query("SELECT id, current_env FROM projects WHERE name = %s", (project_name,))
        if not proj:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        project_id = proj[0]['id']
        current_env = proj[0].get('current_env', '') or ''

        # 检查环境是否存在
        env = db_handler.query("SELECT name FROM environments WHERE project_id = %s AND name = %s", (project_id, env_name))
        if not env:
            return jsonify({'success': False, 'error': f'环境不存在: {env_name}'})

        # 删除环境
        db_handler.execute("DELETE FROM environments WHERE project_id = %s AND name = %s", (project_id, env_name))

        # 如果删除的是当前环境，切换到第一个可用环境
        if current_env == env_name:
            remaining = db_handler.query("SELECT name FROM environments WHERE project_id = %s ORDER BY id LIMIT 1", (project_id,))
            new_env = remaining[0]['name'] if remaining else ''
            db_handler.execute("UPDATE projects SET current_env = %s WHERE id = %s", (new_env, project_id))

        return jsonify({'success': True, 'message': f'环境 {env_name} 已删除'})
    except Exception as e:
        logger.error(f"删除环境失败: {e}")
        return jsonify({'success': False, 'error': f'删除失败: {str(e)}'})


# 路由：获取项目变量列表
@app.route('/projects/<project_name>/variables/list')
def project_variables_list(project_name):
    """获取项目的自定义变量列表"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    projects_data = get_projects()
    if project_name not in projects_data:
        return jsonify({'success': False, 'error': '项目不存在'}), 404

    variables = projects_data[project_name].get('variables', {})
    var_list = [{'key': k, 'value': v} for k, v in variables.items()]
    return jsonify({'success': True, 'variables': var_list})


# 路由：保存项目变量
@app.route('/projects/<project_name>/variables/save', methods=['POST'])
def project_variables_save(project_name):
    """添加或编辑项目变量"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    var_key = request.form.get('var_key', '').strip()
    var_value = request.form.get('var_value', '').strip()

    if not var_key:
        return jsonify({'success': False, 'error': '变量名不能为空'})

    try:
        db_handler._ensure_connection()
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        project_id = proj[0]['id']

        # 插入或更新变量（利用UNIQUE KEY uk_project_var）
        db_handler.execute(
            "INSERT INTO variables (project_id, name, value) VALUES (%s, %s, %s) "
            "ON DUPLICATE KEY UPDATE value = %s",
            (project_id, var_key, var_value, var_value)
        )

        return jsonify({'success': True, 'message': f'变量 {var_key} 保存成功'})
    except Exception as e:
        logger.error(f"保存变量失败: {e}")
        return jsonify({'success': False, 'error': f'保存失败: {str(e)}'})


# 路由：删除项目变量
@app.route('/projects/<project_name>/variables/delete/<var_key>', methods=['POST'])
def project_variables_delete(project_name, var_key):
    """删除项目变量"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    var_key = unquote(var_key)

    try:
        db_handler._ensure_connection()
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        project_id = proj[0]['id']

        var = db_handler.query("SELECT name FROM variables WHERE project_id = %s AND name = %s", (project_id, var_key))
        if not var:
            return jsonify({'success': False, 'error': f'变量不存在: {var_key}'})

        db_handler.execute("DELETE FROM variables WHERE project_id = %s AND name = %s", (project_id, var_key))

        return jsonify({'success': True, 'message': f'变量 {var_key} 已删除'})
    except Exception as e:
        logger.error(f"删除变量失败: {e}")
        return jsonify({'success': False, 'error': f'删除失败: {str(e)}'})


# 路由：批量保存项目变量
@app.route('/projects/<project_name>/variables/batch_save', methods=['POST'])
def project_variables_batch_save(project_name):
    """批量保存项目变量"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    variables_json = request.form.get('variables', '{}')

    try:
        variables = json.loads(variables_json)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'success': False, 'error': '变量数据格式错误'})

    if not isinstance(variables, dict):
        return jsonify({'success': False, 'error': '变量数据格式错误'})

    try:
        db_handler._ensure_connection()
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        project_id = proj[0]['id']

        # 先删除旧变量，再批量插入新变量
        db_handler.execute("DELETE FROM variables WHERE project_id = %s", (project_id,))
        for key, value in variables.items():
            db_handler.execute(
                "INSERT INTO variables (project_id, name, value) VALUES (%s, %s, %s)",
                (project_id, key, value)
            )

        return jsonify({'success': True, 'message': '变量保存成功'})
    except Exception as e:
        logger.error(f"批量保存变量失败: {e}")
        return jsonify({'success': False, 'error': f'保存失败: {str(e)}'})


# 路由：批量执行模块下所有测试
@app.route('/test/execute_module/<project_name>/<module_name>')
def test_execute_module(project_name, module_name):
    """批量执行模块下所有接口测试"""
    from urllib.parse import unquote
    project_name = unquote(project_name)
    module_name = unquote(module_name)

    # 从数据库中获取模块下的所有接口
    apis = []
    if db_handler:
        try:
            db_handler._ensure_connection()
            # 查询项目ID
            proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
            if proj:
                proj_id = proj[0]['id']
                # 查询模块ID
                mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
                if mod:
                    mod_id = mod[0]['id']
                    # 查询模块下的所有接口，按定时任务排序顺序执行
                    api_list = db_handler.query(
                        """SELECT a.* FROM apis a
                           LEFT JOIN scheduler_jobs sj ON sj.case_index = a.id AND sj.project_name = %s AND sj.module_name = %s
                           WHERE a.module_id = %s
                           ORDER BY COALESCE(sj.sort_order, 999), a.id""",
                        (project_name, module_name, mod_id)
                    )
                    if api_list:
                        for api in api_list:
                            # 构造测试数据，与定时任务的处理方式一致
                            api_data = {
                                'api_id': api.get('id'),
                                'case_name': api.get('case_name', ''),
                                'url': api.get('url', ''),
                                'method': api.get('method', 'GET'),
                                'headers': json.loads(api.get('headers', '{}')),
                                'data': _unwrap_raw(json.loads(api.get('data', '{}'))),
                                'expected': json.loads(api.get('expected', '{}')),
                                'extractions': json.loads(api.get('extractions', '{}')),
                                'project': project_name,
                                'module': module_name,
                                'source': '执行'
                            }
                            apis.append(api_data)
        except Exception as e:
            logger.error(f"获取模块接口数据失败: {e}", exc_info=True)

    if not apis:
        return jsonify({'success': False, 'error': '模块不存在或没有接口', 'results': []})

    # 逐个执行测试
    results = []
    passed = 0
    failed = 0

    # 变量由execute_api_test内部从数据库读取，无需外部传递

    for i, api_data in enumerate(apis):
        # api_data已经从数据库中构造，包含所有必要信息 
        try:
            # 变量由execute_api_test内部从数据库读取最新值
            result = execute_api_test(api_data)

            # 变量已由execute_api_test提取并保存到数据库，下一个接口自动从数据库读取最新值
            if result.get('extraction_results'):
                for extraction in result['extraction_results']:
                    if extraction.get('success') and extraction.get('var_name'):
                        logger.info(f"[运行全部] 接口提取变量已保存到数据库: {extraction['var_name']} = {extraction.get('value')}")

            result['case_name'] = api_data.get('case_name', f'接口{i+1}')
            result['index'] = i
            result['api_id'] = api_data.get('api_id')
            if result.get('success'):
                passed += 1
            else:
                failed += 1
            results.append(result)
        except Exception as e:
            failed += 1
            results.append({
                'success': False,
                'error': str(e),
                'case_name': api_data.get('case_name', f'接口{i+1}'),
                'index': i,
                'api_id': api_data.get('api_id')
            })

    return jsonify({
        'success': True,
        'total': len(apis),
        'passed': passed,
        'failed': failed,
        'results': results
    })


# 路由：执行测试
@app.route('/test/execute/<project_name>/<module_name>/<int:api_id>')
def test_execute(project_name, module_name, api_id):
    """执行测试"""
    logger.info(f"========== 开始执行测试 ==========")
    logger.info(f"项目名称: {project_name}, 模块名称: {module_name}, 接口ID: {api_id}")

    # 对URL编码的参数进行解码
    from urllib.parse import unquote
    project_name = unquote(project_name)
    module_name = unquote(module_name)

    if not db_handler:
        logger.error("数据库未连接")
        return jsonify({"error": "数据库未连接", "success": False})

    try:
        db_handler._ensure_connection()
        logger.info("数据库连接正常")

        # 查询项目ID
        proj = db_handler.query("SELECT id FROM projects WHERE name = %s", (project_name,))
        if not proj:
            logger.warning(f"项目不存在: {project_name}")
            return jsonify({"error": "项目不存在", "success": False})
        proj_id = proj[0]['id']

        # 查询模块ID
        mod = db_handler.query("SELECT id FROM modules WHERE project_id = %s AND name = %s", (proj_id, module_name))
        if not mod:
            logger.warning(f"模块不存在: {module_name}")
            return jsonify({"error": "模块不存在", "success": False})
        mod_id = mod[0]['id']

        # 查询接口
        api = db_handler.query("SELECT * FROM apis WHERE id = %s", (api_id,))
        if not api:
            logger.warning(f"接口不存在: {api_id}")
            return jsonify({"error": "接口不存在", "success": False})

        api_data = api[0]
        logger.info(f"获取接口数据: {api_data}")

        # 构造测试数据
        test_data = {
            'case_name': api_data.get('case_name', ''),
            'url': api_data.get('url', ''),
            'method': api_data.get('method', 'GET'),
            'headers': json.loads(api_data.get('headers', '{}')),
            'data': _unwrap_raw(json.loads(api_data.get('data', '{}'))),
            'expected': json.loads(api_data.get('expected', '{}')),
            'extractions': json.loads(api_data.get('extractions', '{}')),
            'project': project_name,
            'module': module_name,
            'source': '执行'
        }

        logger.info(f"测试数据: {test_data}")

        # 变量替换和提取均在execute_api_test内部完成，每次执行前从数据库读取最新变量
        # 执行测试
        result = execute_api_test(test_data)
        return jsonify(result)
    except Exception as e:
        import traceback
        tb_str = traceback.format_exc()
        logger.error(f"test_execute路由异常: {e}\n{tb_str}")
        # 将traceback写入调试文件
        try:
            debug_path = os.path.join(BASE_DIR, 'debug_error.log')
            with open(debug_path, 'w', encoding='utf-8') as f:
                f.write(f"=== test_execute路由异常 ===\n")
                f.write(f"error: {e}\n")
                f.write(f"traceback:\n{tb_str}\n")
        except Exception:
            pass
        return jsonify({"error": f"{e}\n{tb_str}", "success": False})


# 路由：从备份恢复统计数据
@app.route('/statistics/restore', methods=['POST'])
def statistics_restore():
    """从备份恢复统计数据"""
    try:
        backup_file = request.form.get('backup_file')
        if not backup_file:
            return jsonify({'success': False, 'message': '未指定备份文件'}), 400
            
        if not os.path.exists(backup_file):
            return jsonify({'success': False, 'message': '备份文件不存在'}), 404
            
        # 读取备份数据
        try:
            with open(backup_file, 'r', encoding='utf-8') as f:
                backup_data = json.load(f)
        except Exception as e:
            logger.error(f"读取备份文件失败: {str(e)}")
            return jsonify({'success': False, 'message': f'读取备份文件失败: {str(e)}'}), 500
            
        # 尝试恢复数据
        saved = save_statistics_data(backup_data)
        if saved:
            return jsonify({'success': True, 'message': '统计数据已从备份恢复'})
        else:
            logger.error("恢复统计数据失败：数据保存失败")
            return jsonify({'success': False, 'message': '恢复统计数据失败'}), 500
    except Exception as e:
        logger.error(f"恢复统计数据时发生异常: {str(e)}")
        return jsonify({'success': False, 'message': f'恢复统计数据失败: {str(e)}'}), 500


# 路由：获取所有备份文件列表
@app.route('/statistics/backup/list', methods=['GET'])
def statistics_backup_list():
    """获取所有备份文件列表"""
    try:
        backup_dir = os.path.dirname(STATISTICS_FILE)
        backup_files = []
        
        # 查找所有备份文件
        for filename in os.listdir(backup_dir):
            if filename.startswith(os.path.basename(STATISTICS_FILE) + ".bak_"):
                filepath = os.path.join(backup_dir, filename)
                file_stat = os.stat(filepath)
                backup_files.append({
                    'filename': filename,
                    'path': filepath,
                    'size': file_stat.st_size,
                    'created': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(file_stat.st_ctime))
                })
        
        # 按创建时间降序排序
        backup_files.sort(key=lambda x: x['created'], reverse=True)
        
        return jsonify({
            'success': True, 
            'backup_files': backup_files
        })
    except Exception as e:
        logger.error(f"获取备份文件列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取备份文件列表失败: {str(e)}'}), 500


# 路由：下载备份文件
@app.route('/statistics/download')
def download_backup():
    """下载备份文件"""
    try:
        file_path = request.args.get('file')
        if not file_path:
            return jsonify({'success': False, 'message': '未指定文件路径'}), 400
            
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': '文件不存在'}), 404
            
        # 安全检查，确保只允许下载备份文件
        if not file_path.startswith(os.path.dirname(STATISTICS_FILE)) or \
           not os.path.basename(file_path).startswith(os.path.basename(STATISTICS_FILE) + ".bak_"):
            return jsonify({'success': False, 'message': '非法文件路径'}), 403
            
        # 获取文件名
        filename = os.path.basename(file_path)
        
        # 设置响应头，触发下载
        from flask import send_file
        return send_file(file_path, as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"下载备份文件失败: {str(e)}")
        return jsonify({'success': False, 'message': f'下载失败: {str(e)}'}), 500


# 启动时同步文件中未入库的记录到数据库
sync_file_records_to_db()


if __name__ == '__main__':
    # 创建必要的目录
    templates_dir = os.path.join(BASE_DIR, "templates")
    os.makedirs(templates_dir, exist_ok=True)

    static_dir = os.path.join(BASE_DIR, "static")
    os.makedirs(static_dir, exist_ok=True)

    css_dir = os.path.join(static_dir, "css")
    os.makedirs(css_dir, exist_ok=True)

    js_dir = os.path.join(static_dir, "js")
    os.makedirs(js_dir, exist_ok=True)

    # 启动应用（生产模式使用Waitress）
    try:
        from waitress import serve
        print("[生产模式] 使用Waitress WSGI服务器")
        serve(app, host='0.0.0.0', port=5000, threads=4)
    except ImportError:
        print("警告: waitress未安装，回退到Flask开发服务器")
        app.run(debug=False, host='0.0.0.0', port=5000)
